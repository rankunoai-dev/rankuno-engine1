"""Tests for the local HTTP API.

No test here performs real network I/O. `PageClassificationTool` is replaced
wholesale where a crawl would otherwise run, so what is under test is the
*transport*: admission control, status transitions, and the shape of what a
polling client receives.

The one thing worth stating up front: the API must never present an unfinished
job as a finished one, and must never leave a job in a state nothing will move
it out of. Most of these tests are about those two properties.
"""

from __future__ import annotations

import io
import time

import httpx
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from src.api import server as server_module
from src.api.server import API_PREFIX, create_app
from src.core.state_store import MAX_HOMEPAGE_BYTES, DiskJobStore, JobRecord, JobStatus
from src.core.url_safety import UrlSafetyPolicy
from src.modules.seo.page_classifier.discovery import DiscoveryReport, SiteGraph
from src.modules.seo.page_classifier.schemas import (
    ConsensusMethod,
    FullPageIntelligenceProfile,
    HierarchyLevel,
    PrimaryPageType,
    SearchIntent,
    SignalScore,
    SignalSource,
)
from src.modules.seo.page_classifier.tool import (
    CrawlSummary,
    PageClassificationOutput,
)
from src.modules.seo.page_classifier.weights import SiteProfile, WeightProfileReport

PUBLIC_IP = "93.184.216.34"
SAFE_URL = "https://e.com/"


class StubResult:
    """Stands in for `ToolResult` without importing the generic machinery."""

    def __init__(self, ok: bool = True, data: object = None, error: str | None = None) -> None:
        """Record what the stubbed tool should report."""
        self.ok = ok
        self.data = data
        self.error = error


class StubTool:
    """A `PageClassificationTool` that returns instantly instead of crawling."""

    result: StubResult = StubResult(ok=False, error="not configured")

    def __init__(self, **_kwargs: object) -> None:
        """Accept and ignore the real tool's constructor arguments."""

    def run(self, _payload: object) -> StubResult:
        return type(self).result


@pytest.fixture
def stub_tool(monkeypatch):
    """Replace the real tool for the duration of a test."""
    monkeypatch.setattr(server_module, "PageClassificationTool", StubTool)
    return StubTool


@pytest.fixture
def store(tmp_path) -> DiskJobStore:
    return DiskJobStore(tmp_path / "jobs")


@pytest.fixture
def client(store):
    """A client whose SSRF resolver is deterministic, not DNS-dependent."""
    app = create_app(
        store=store,
        url_policy=UrlSafetyPolicy(resolver=lambda host: [PUBLIC_IP]),
    )
    with TestClient(app) as test_client:
        yield test_client


def post_job(client, url: str = SAFE_URL, **overrides: object):
    body = {"base_url": url, "max_pages": 5, "crawl_dom": False, **overrides}
    return client.post(f"{API_PREFIX}/jobs", json=body)


def run_job(client, store, url: str = SAFE_URL, **overrides: object) -> str:
    """Submit a job and wait for the worker to finish it.

    The wait is not test scaffolding around a slow system — it is the contract.
    `POST /jobs` returns `202` the moment the work is *accepted*, so a job is
    still `queued` or `running` when the response arrives. Asserting on its
    outcome immediately would be asserting on a race.

    Returns:
        The job id.

    Raises:
        AssertionError: If the job does not finish, which means the worker was
            never dispatched or died without recording a terminal status.
    """
    response = post_job(client, url, **overrides)
    assert response.status_code == 202, response.text
    job_id: str = response.json()["id"]

    deadline = time.monotonic() + 5.0
    while time.monotonic() < deadline:
        if store.get(job_id).is_terminal:
            return job_id
        time.sleep(0.01)

    pytest.fail(f"job {job_id} never reached a terminal status")


class TestHealth:
    def test_reports_ok(self, client):
        response = client.get(f"{API_PREFIX}/health")
        assert response.status_code == 200
        assert response.json()["status"] == "ok"

    def test_reports_spare_capacity(self, client):
        body = client.get(f"{API_PREFIX}/health").json()
        assert body["active_jobs"] == 0
        assert body["max_concurrent_jobs"] >= 1


class TestAdmissionControl:
    def test_a_safe_url_is_accepted_with_202(self, client, stub_tool):
        """202, not 200: nothing has been crawled when this returns."""
        stub_tool.result = StubResult(ok=False, error="stopped")
        response = post_job(client)
        assert response.status_code == 202
        assert response.json()["id"]

    @pytest.mark.parametrize(
        "hostile",
        [
            "http://127.0.0.1:8000/",
            "http://localhost/",
            "http://[::1]/",
            "http://169.254.169.254/latest/meta-data/",
            "file:///etc/passwd",
        ],
    )
    def test_ssrf_targets_are_rejected_at_admission(self, client, hostile):
        """The reason this check exists at the door, not only in the crawl.

        `169.254.169.254` is the cloud metadata endpoint — the single most
        valuable SSRF target there is. A rejection must be a 400 the operator
        sees, not a job that fails quietly some milliseconds later.
        """
        assert post_job(client, hostile).status_code == 400

    def test_a_rejected_url_creates_no_job(self, client, store):
        post_job(client, "http://127.0.0.1/")
        assert store.list_jobs() == []

    def test_the_rejection_says_why(self, client):
        detail = post_job(client, "http://127.0.0.1/").json()["detail"]
        assert detail, "a 400 with no reason is unactionable"

    def test_a_malformed_body_is_a_422(self, client):
        assert client.post(f"{API_PREFIX}/jobs", json={"max_pages": 5}).status_code == 422

    def test_unknown_fields_are_rejected(self, client):
        """`StrictModel` forbids extras, so a typo'd field cannot be ignored."""
        response = post_job(client, surprise=1)
        assert response.status_code == 422


class TestConcurrencyCap:
    def test_excess_jobs_are_refused_with_429(self, store):
        """Each in-flight crawl holds its whole graph in memory."""
        app = create_app(
            store=store,
            url_policy=UrlSafetyPolicy(resolver=lambda host: [PUBLIC_IP]),
            max_concurrent_jobs=1,
        )
        state = app.state.api
        assert state.try_reserve("occupier") is True

        with TestClient(app) as client:
            response = post_job(client)
        assert response.status_code == 429

    def test_a_refused_job_leaves_no_record(self, store):
        """A refusal is not a job, and must not look like one afterwards.

        This asserted the opposite until the ordering was fixed: capacity was
        checked *after* the store minted a record, so every refusal persisted a
        job and immediately marked it failed. Retrying against a full server
        manufactured one permanent FAILED row per click — 16 of 99 records on
        one workstation were nothing but refusals, and in the jobs list they
        were indistinguishable from crawls that had really run and really
        failed.

        Nothing waits forever as a result: the 429 is synchronous, so the caller
        is told at the point of asking and never receives an id to poll.
        """
        app = create_app(
            store=store,
            url_policy=UrlSafetyPolicy(resolver=lambda host: [PUBLIC_IP]),
            max_concurrent_jobs=1,
        )
        app.state.api.try_reserve("occupier")
        with TestClient(app) as client:
            assert post_job(client).status_code == 429

        assert store.list_jobs() == []

    def test_a_refusal_does_not_consume_a_slot(self, store):
        """The provisional reservation must be given back, not leaked.

        Capacity is now claimed before the record exists, so a refused request
        holds a slot for the duration of the check. Leaking one would cost a
        permanent slot per refused click — the exact failure this whole change
        is about.
        """
        app = create_app(
            store=store,
            url_policy=UrlSafetyPolicy(resolver=lambda host: [PUBLIC_IP]),
            max_concurrent_jobs=1,
        )
        state = app.state.api
        state.try_reserve("occupier")
        with TestClient(app) as client:
            for _ in range(5):
                assert post_job(client).status_code == 429

        state.release("occupier")
        assert state.active_count == 0

    def test_reserving_is_atomic(self, store):
        """Check-then-reserve in two steps would let both callers through."""
        app = create_app(store=store, max_concurrent_jobs=2)
        state = app.state.api
        assert [state.try_reserve(f"j{index}") for index in range(3)] == [True, True, False]

    def test_releasing_frees_a_slot(self, store):
        app = create_app(store=store, max_concurrent_jobs=1)
        state = app.state.api
        state.try_reserve("a")
        assert state.try_reserve("b") is False
        state.release("a")
        assert state.try_reserve("b") is True

    def test_releasing_an_unknown_job_is_harmless(self, store):
        """`_dispatch` releases in a `finally`; a double release must not raise."""
        state = create_app(store=store).state.api
        state.release("never-reserved")
        assert state.active_count == 0


class TestJobLifecycle:
    def test_a_successful_crawl_reaches_succeeded(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data=_fake_output(truncated=False))
        job_id = run_job(client, store)
        assert store.get(job_id).status is JobStatus.SUCCEEDED

    def test_a_truncated_crawl_reaches_partial(self, client, store, stub_tool):
        """The UI must be able to say the crawl is incomplete."""
        stub_tool.result = StubResult(ok=True, data=_fake_output(truncated=True))
        job_id = run_job(client, store)
        assert store.get(job_id).status is JobStatus.PARTIAL

    def test_a_tool_failure_reaches_failed_with_a_reason(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=False, error="robots.txt disallowed /")
        job_id = run_job(client, store)
        record = store.get(job_id)
        assert record.status is JobStatus.FAILED
        assert record.error == "robots.txt disallowed /"

    def test_a_crashing_tool_still_reaches_a_terminal_state(self, client, store, monkeypatch):
        """The worker runs detached. An escaping exception would strand the job.

        It would stay `RUNNING` with nothing left to move it, and a polling UI
        would wait indefinitely for a result that is never coming.
        """

        class ExplodingTool:
            def __init__(self, **_kwargs: object) -> None:
                pass

            def run(self, _payload: object) -> object:
                raise RuntimeError("transport exploded")

        monkeypatch.setattr(server_module, "PageClassificationTool", ExplodingTool)
        job_id = run_job(client, store)

        record = store.get(job_id)
        assert record.status is JobStatus.FAILED
        assert "transport exploded" in (record.error or "")

    def test_a_finished_job_frees_its_slot(self, client, store, stub_tool):
        """A slot leaked on completion would cap the server at three crawls ever."""
        stub_tool.result = StubResult(ok=True, data=_fake_output(truncated=False))
        run_job(client, store)
        assert client.get(f"{API_PREFIX}/health").json()["active_jobs"] == 0


class TestReadingJobs:
    def test_an_unknown_job_is_404(self, client):
        assert client.get(f"{API_PREFIX}/jobs/nope").status_code == 404

    def test_a_missing_result_is_409_not_404(self, client, store):
        """404 would tell a polling client to give up on a live job.

        The record is created directly rather than through `POST`, so the job
        stays `running` instead of racing a worker to a terminal status.
        """
        job_id = store.create("seo.page_classifier", {"base_url": SAFE_URL}).id
        store.mark_running(job_id)

        response = client.get(f"{API_PREFIX}/jobs/{job_id}/result")
        assert response.status_code == 409

    def test_a_result_is_returned_once_finished(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data=_fake_output(truncated=False))
        job_id = run_job(client, store)

        response = client.get(f"{API_PREFIX}/jobs/{job_id}/result")
        assert response.status_code == 200
        assert response.json()["base_url"] == SAFE_URL

    def test_the_job_list_omits_result_blobs(self, client, store, stub_tool):
        """Listing must stay cheap; a 20k result is ~16 MB."""
        stub_tool.result = StubResult(ok=True, data=_fake_output(truncated=False))
        run_job(client, store)

        listing = client.get(f"{API_PREFIX}/jobs").json()
        assert len(listing) == 1
        assert "result" not in listing[0]
        assert listing[0]["has_result"] is True

    def test_the_status_payload_carries_what_a_poller_needs(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data=_fake_output(truncated=False))
        job_id = run_job(client, store)

        body = client.get(f"{API_PREFIX}/jobs/{job_id}").json()
        assert body["status"] == "succeeded"
        assert body["has_result"] is True
        assert body["label"] == SAFE_URL


class TestCors:
    def test_the_vite_origin_is_allowed(self, client):
        response = client.get(f"{API_PREFIX}/health", headers={"Origin": "http://localhost:5173"})
        assert response.headers.get("access-control-allow-origin") == "http://localhost:5173"

    def test_an_arbitrary_origin_is_not_allowed(self, client):
        """A wildcard would let any page the operator visits read crawl results."""
        response = client.get(f"{API_PREFIX}/health", headers={"Origin": "https://evil.test"})
        assert response.headers.get("access-control-allow-origin") is None


class TestStartupRecovery:
    def test_orphaned_jobs_are_failed_on_startup(self, tmp_path):
        """A job left RUNNING by a killed process has no worker any more."""
        store = DiskJobStore(tmp_path / "jobs")
        job_id = store.create("seo.page_classifier", {"base_url": SAFE_URL}).id
        store.mark_running(job_id)

        with TestClient(create_app(store=store)):
            pass

        assert store.get(job_id).status is JobStatus.FAILED


def _fake_output(*, truncated: bool) -> PageClassificationOutput:
    """An empty but genuine `PageClassificationOutput`.

    A real instance rather than a stand-in, because the server type-checks what
    the tool returned before persisting it. A duck-typed object would pass the
    tests while the production path rejected the real thing.
    """
    return PageClassificationOutput(
        base_url=SAFE_URL,
        site_profile=SiteProfile(),
        weight_profile=WeightProfileReport.for_site(SiteProfile()),
        discovery=DiscoveryReport(base_url=SAFE_URL, truncated=truncated),
        summary=CrawlSummary(),
    )


class TestBlockedCrawlSurfacing:
    """A blocked crawl must reach the client as a failure, not as a result.

    The tool raises `CrawlBlockedError`, `BaseTool.run()` converts it to a
    non-success `ToolResult`, and the API must persist that as `FAILED` with the
    reason intact. If any link in that chain drops the message, the UI shows a
    green job with one confident page and no explanation.
    """

    def test_a_blocked_crawl_becomes_a_failed_job(self, client, store, stub_tool):
        stub_tool.result = StubResult(
            ok=False,
            error="Crawl failed: all 6 requests were refused by the target server.",
        )
        job_id = run_job(client, store)

        record = store.get(job_id)
        assert record.status is JobStatus.FAILED
        assert record.has_result is False, "no result blob for a crawl that produced nothing"

    def test_the_refusal_reason_reaches_the_client(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=False, error="all 6 requests were refused")
        job_id = run_job(client, store)

        body = client.get(f"{API_PREFIX}/jobs/{job_id}").json()
        assert "refused" in (body["error"] or "")

    def test_a_failed_job_offers_no_result_endpoint(self, client, store, stub_tool):
        """409, so a poller learns it has stopped rather than retrying forever."""
        stub_tool.result = StubResult(ok=False, error="all requests refused")
        job_id = run_job(client, store)

        assert client.get(f"{API_PREFIX}/jobs/{job_id}/result").status_code == 409


class TestTelemetryRecorder:
    """Progress reporting must not cost more than the work it reports on.

    Each write rewrites the job record through `os.replace` and an `fsync`, so
    an unthrottled recorder on a 20,000-page crawl would spend more wall-clock
    in the filesystem than on the network.
    """

    def test_the_first_call_flushes_immediately(self, store):
        """A crawl that shows nothing until half a second in looks stuck."""
        job_id = store.create("seo.page_classifier", {}).id
        server_module.TelemetryRecorder(store, job_id, 100)(0, 50, ())
        assert store.get(job_id).telemetry.discovered == 50

    def test_rapid_calls_are_throttled(self, store):
        job_id = store.create("seo.page_classifier", {}).id
        recorder = server_module.TelemetryRecorder(store, job_id, 100)
        recorder(0, 50, ())
        for completed in range(1, 60):
            recorder(completed, 50, ())
        # Only the initial flush should have landed inside half a second.
        assert store.get(job_id).telemetry.completed == 0

    def test_progress_is_measured_against_discovery_not_the_ceiling(self, store):
        """A 300-page site crawled with a 20,000 ceiling finishes at 300.

        Dividing by the ceiling would leave the bar at 1.5% forever.
        """
        job_id = store.create("seo.page_classifier", {}).id
        server_module.TelemetryRecorder(store, job_id, 20_000)(300, 300, ())
        telemetry = store.get(job_id).telemetry
        assert telemetry.discovered == 300
        assert telemetry.fraction == 1.0

    def test_discovery_beyond_the_ceiling_is_capped(self, store):
        """The crawl stops at the ceiling, so the denominator does too."""
        job_id = store.create("seo.page_classifier", {}).id
        server_module.TelemetryRecorder(store, job_id, 500)(0, 4_427, ())
        assert store.get(job_id).telemetry.discovered == 500

    def test_no_eta_during_warmup(self, store):
        """A rate over the first fraction of a second swings by orders of magnitude."""
        job_id = store.create("seo.page_classifier", {}).id
        server_module.TelemetryRecorder(store, job_id, 100)(5, 100, ())
        assert store.get(job_id).telemetry.eta_seconds is None

    def test_recent_items_are_capped(self, store):
        """20,000 URLs on every poll would cost more than the crawl."""
        job_id = store.create("seo.page_classifier", {}).id
        urls = tuple(f"https://e.com/{n}/" for n in range(200))
        server_module.TelemetryRecorder(store, job_id, 1_000)(0, 200, urls)

        recent = store.get(job_id).telemetry.recent_items
        assert len(recent) == 20
        assert recent[-1] == "https://e.com/199/", "newest last"

    def test_a_deleted_job_does_not_break_the_crawl(self, store):
        """Losing telemetry must not interrupt work still producing a result."""
        recorder = server_module.TelemetryRecorder(store, "gone", 100)
        recorder(1, 10, ())  # must not raise

    def test_telemetry_survives_into_the_status_payload(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data=_fake_output(truncated=False))
        job_id = run_job(client, store)
        assert "telemetry" in client.get(f"{API_PREFIX}/jobs/{job_id}").json()


class TestTelemetryContract:
    def test_fraction_is_none_before_anything_is_discovered(self):
        """`None`, not zero: an unknown total is not a total of zero."""
        from src.core.state_store import JobTelemetry

        assert JobTelemetry().fraction is None

    def test_fraction_cannot_exceed_one(self):
        """Discovery shrinks when duplicates merge; the bar must not overrun."""
        from src.core.state_store import JobTelemetry

        assert JobTelemetry(completed=120, discovered=100).fraction == 1.0


class TestCheckpointRecovery:
    """A crawl killed by the process dying must not lose what it found.

    Cycle 0018 covers in-process failures: a stall or an exception still returns
    a partial result. This covers the case that one cannot, where the process is
    gone and whatever reached disk is all there is.
    """

    def test_the_first_write_is_immediate(self, store):
        graph = SiteGraph("https://e.com", max_pages=100)
        graph.add("https://e.com/a/", sitemap=True)
        job_id = store.create("seo.page_classifier", {}).id

        server_module.CrawlCheckpointer(store, job_id, "https://e.com")(graph)
        assert store.get(job_id).has_checkpoint is True

    def test_rapid_calls_are_throttled(self, store):
        """20,000 URLs through fsync per page would cost more than the crawl."""
        graph = SiteGraph("https://e.com", max_pages=1_000)
        graph.add("https://e.com/a/", sitemap=True)
        job_id = store.create("seo.page_classifier", {}).id
        checkpointer = server_module.CrawlCheckpointer(store, job_id, "https://e.com")
        checkpointer(graph)

        for index in range(50):
            graph.add(f"https://e.com/p{index}/", sitemap=True)
            checkpointer(graph)

        saved = store.read_checkpoint(job_id)
        assert saved is not None
        assert len(saved["urls"]) == 1, "only the first write landed inside the window"

    def test_a_page_count_boundary_forces_a_write(self, store):
        """Time alone under-saves a fast crawl: Turbo puts 250 pages in 10s."""
        graph = SiteGraph("https://e.com", max_pages=1_000)
        job_id = store.create("seo.page_classifier", {}).id
        checkpointer = server_module.CrawlCheckpointer(store, job_id, "https://e.com")

        for index in range(server_module.CHECKPOINT_EVERY_PAGES + 5):
            graph.add(f"https://e.com/p{index}/", sitemap=True)
            checkpointer(graph)

        saved = store.read_checkpoint(job_id)
        assert saved is not None
        assert len(saved["urls"]) >= server_module.CHECKPOINT_EVERY_PAGES

    def test_an_empty_graph_writes_nothing(self, store):
        """A checkpoint of zero URLs is not worth an fsync."""
        job_id = store.create("seo.page_classifier", {}).id
        server_module.CrawlCheckpointer(store, job_id, "https://e.com")(
            SiteGraph("https://e.com", max_pages=10)
        )
        assert store.get(job_id).has_checkpoint is False

    def test_a_deleted_job_does_not_break_the_crawl(self, store):
        graph = SiteGraph("https://e.com", max_pages=10)
        graph.add("https://e.com/a/", sitemap=True)
        server_module.CrawlCheckpointer(store, "gone", "https://e.com")(graph)

    def test_a_corrupt_checkpoint_reads_as_absent(self, store):
        """A truncated write is what a power cut produces; it must not raise."""
        job_id = store.create("seo.page_classifier", {}).id
        (store.root / f"{job_id}.checkpoint.json").write_text("{trunca", encoding="utf-8")
        assert store.read_checkpoint(job_id) is None


class TestCheckpointEndpoint:
    def test_a_saved_checkpoint_renders_as_an_output(self, client, store):
        job_id = store.create("seo.page_classifier", {}).id
        store.write_checkpoint(
            job_id, {"base_url": SAFE_URL, "urls": [f"{SAFE_URL}p{n}/" for n in range(5)]}
        )

        body = client.get(f"{API_PREFIX}/jobs/{job_id}/checkpoint").json()
        assert body["base_url"] == SAFE_URL
        assert len(body["pages"]) == 5

    def test_recovered_pages_are_marked_unclassified(self, client, store):
        """A checkpoint holds URLs, not classifications. Saying otherwise lies."""
        job_id = store.create("seo.page_classifier", {}).id
        store.write_checkpoint(job_id, {"base_url": SAFE_URL, "urls": [f"{SAFE_URL}a/"]})

        body = client.get(f"{API_PREFIX}/jobs/{job_id}/checkpoint").json()
        assert body["pages"][0]["primary_page_type"] == "UNKNOWN"
        assert body["pages"][0]["final_confidence_score"] == 0.0
        assert body["summary"]["unknown_pages"] == 1

    def test_the_recovered_view_says_it_is_recovered(self, client, store):
        """It must never be mistaken for a completed crawl."""
        job_id = store.create("seo.page_classifier", {}).id
        store.write_checkpoint(job_id, {"base_url": SAFE_URL, "urls": [f"{SAFE_URL}a/"]})

        body = client.get(f"{API_PREFIX}/jobs/{job_id}/checkpoint").json()
        assert "interrupted" in body["discovery"]["stopped_reason"]

    def test_a_job_with_no_checkpoint_is_404(self, client, store):
        job_id = store.create("seo.page_classifier", {}).id
        assert client.get(f"{API_PREFIX}/jobs/{job_id}/checkpoint").status_code == 404

    def test_an_unknown_job_is_404(self, client):
        assert client.get(f"{API_PREFIX}/jobs/nope/checkpoint").status_code == 404


class TestRetry:
    """Re-running a job's crawl with the settings it originally used.

    A new job every time, never a mutation of the old one: the original record
    is the evidence of what happened and when, and a failed crawl is frequently
    the finding itself.
    """

    def test_starts_a_new_job(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)

        response = client.post(f"{API_PREFIX}/jobs/{original}/retry")
        assert response.status_code == 202, response.text
        assert response.json()["id"] != original

    def test_the_original_job_is_untouched(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        before = store.get(original)

        client.post(f"{API_PREFIX}/jobs/{original}/retry")

        after = store.get(original)
        assert after.status == before.status
        assert after.finished_at == before.finished_at

    def test_it_reuses_the_original_settings(self, client, store, stub_tool):
        """Not the defaults. A retry at different settings is a different crawl."""
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store, max_pages=7, crawl_dom=False, concurrency=2)

        retried = client.post(f"{API_PREFIX}/jobs/{original}/retry").json()["id"]

        request = store.get(retried).request
        assert request["max_pages"] == 7
        assert request["concurrency"] == 2
        assert request["crawl_dom"] is False

    def test_a_failed_job_can_be_retried(self, client, store, stub_tool):
        """The case this exists for: openai.com returning 403."""
        stub_tool.result = StubResult(ok=False, error="403 from the target")
        original = run_job(client, store)
        assert store.get(original).status is JobStatus.FAILED

        assert client.post(f"{API_PREFIX}/jobs/{original}/retry").status_code == 202

    def test_a_successful_job_can_also_be_retried(self, client, store, stub_tool):
        """Re-crawling to pick up site changes is as legitimate as retrying."""
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)

        assert client.post(f"{API_PREFIX}/jobs/{original}/retry").status_code == 202

    def test_an_unknown_job_is_404(self, client):
        assert client.post(f"{API_PREFIX}/jobs/nope/retry").status_code == 404

    def test_a_job_with_no_stored_request_is_409(self, client, store):
        """Records written before the request was kept, or by another tool."""
        record = store.create("some.other.tool", {}, label="legacy")
        response = client.post(f"{API_PREFIX}/jobs/{record.id}/retry")
        assert response.status_code == 409
        assert "cannot be re-run" in response.json()["detail"]

    def test_settings_this_build_rejects_are_409_not_500(self, client, store):
        """A schema change must not reach the operator as a traceback.

        A record written before the change is refused with an explanation
        instead of a pydantic error nobody can act on.
        """
        record = store.create(server_module.TOOL_NAME, {"base_url": SAFE_URL, "max_pages": -5})
        response = client.post(f"{API_PREFIX}/jobs/{record.id}/retry")
        assert response.status_code == 409
        assert "no longer accepts" in response.json()["detail"]

    def test_the_url_is_revalidated_at_retry(self, client, store):
        """A stored payload is not automatically still safe.

        DNS moves. A host that resolved publicly when the job first ran may
        resolve to a private address today, and replaying the payload without
        re-checking would turn a stored record into an SSRF primitive.
        """
        record = store.create(
            server_module.TOOL_NAME, {"base_url": "http://127.0.0.1/", "max_pages": 5}
        )
        response = client.post(f"{API_PREFIX}/jobs/{record.id}/retry")
        assert response.status_code == 400


class TestResume:
    """Crawling the URLs an interrupted job discovered but never fetched."""

    def _checkpoint(self, store, job_id: str, unfetched: list[str]) -> None:
        store.write_checkpoint(
            job_id,
            {
                "base_url": SAFE_URL,
                "urls": [SAFE_URL, *unfetched],
                "unfetched": unfetched,
                "saved_at_count": 1 + len(unfetched),
            },
        )

    def test_seeds_the_new_crawl_with_what_was_missed(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        self._checkpoint(store, original, [f"{SAFE_URL}a", f"{SAFE_URL}b"])

        response = client.post(f"{API_PREFIX}/jobs/{original}/resume")
        assert response.status_code == 202, response.text

        request = store.get(response.json()["id"]).request
        assert request["seed_urls"] == [f"{SAFE_URL}a", f"{SAFE_URL}b"]

    def test_it_excludes_what_the_first_run_already_fetched(self, client, store, stub_tool):
        """Seeding alone was never a resume.

        `seed_urls` adds to the frontier and removes nothing from it, so the
        crawl still began at the site root, followed every link out of it and
        re-fetched the whole site. Live on gep.com: a resume advertising
        "+2,940" rediscovered 5,311 URLs and started from zero.

        The exclusion is the difference, and it is derived — everything the
        checkpoint discovered, minus what it had still to fetch.
        """
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        self._checkpoint(store, original, [f"{SAFE_URL}a"])

        response = client.post(f"{API_PREFIX}/jobs/{original}/resume")
        assert response.status_code == 202, response.text

        request = store.get(response.json()["id"]).request
        assert request["seed_urls"] == [f"{SAFE_URL}a"]
        # The root was discovered and not outstanding, so it was fetched — and
        # excluding it is what stops the traversal restarting from the homepage.
        assert request["exclude_urls"] == [SAFE_URL]

    def test_a_checkpoint_with_no_discovered_set_still_resumes(self, client, store, stub_tool):
        """Old checkpoints predate `urls` and cannot say what was fetched.

        Refusing would throw away real work; resuming without an exclusion
        degrades to the old full re-crawl. It resumes, with an empty exclusion,
        and the engine logs that it could not narrow the run.
        """
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        store.write_checkpoint(
            original,
            {"base_url": SAFE_URL, "unfetched": [f"{SAFE_URL}a"], "saved_at_count": 1},
        )

        response = client.post(f"{API_PREFIX}/jobs/{original}/resume")
        assert response.status_code == 202, response.text
        request = store.get(response.json()["id"]).request
        assert request["seed_urls"] == [f"{SAFE_URL}a"]
        assert request["exclude_urls"] == []

    def test_the_label_states_how_much_is_left(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        self._checkpoint(store, original, [f"{SAFE_URL}{n}" for n in range(3)])

        label = client.post(f"{API_PREFIX}/jobs/{original}/resume").json()["label"]
        assert "resumed +3" in label

    def test_a_job_with_no_checkpoint_is_404(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        response = client.post(f"{API_PREFIX}/jobs/{original}/resume")
        assert response.status_code == 404

    def test_a_checkpoint_predating_the_feature_is_409(self, client, store, stub_tool):
        """A checkpoint without an unfetched set cannot be resumed.

        The discovered set alone cannot separate fetched from unfetched, so
        resuming would silently re-fetch everything — which is `retry`.
        """
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        store.write_checkpoint(
            original,
            {"base_url": SAFE_URL, "urls": [SAFE_URL], "saved_at_count": 1},
        )

        response = client.post(f"{API_PREFIX}/jobs/{original}/resume")
        assert response.status_code == 409
        assert "retry it instead" in response.json()["detail"]

    def test_nothing_left_to_fetch_is_409(self, client, store, stub_tool):
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        self._checkpoint(store, original, [])

        response = client.post(f"{API_PREFIX}/jobs/{original}/resume")
        assert response.status_code == 409
        assert "every URL it discovered" in response.json()["detail"]

    def test_a_running_job_cannot_be_resumed(self, client, store):
        """A running job offers no stable delta.

        Its checkpoint is still moving, so anything read now is stale before the
        resumed crawl starts.
        """
        record = store.create(server_module.TOOL_NAME, {"base_url": SAFE_URL, "max_pages": 5})
        store.mark_running(record.id)
        self._checkpoint(store, record.id, [f"{SAFE_URL}a"])

        response = client.post(f"{API_PREFIX}/jobs/{record.id}/resume")
        assert response.status_code == 409
        assert "still running" in response.json()["detail"]

    def test_an_unknown_job_is_404(self, client):
        assert client.post(f"{API_PREFIX}/jobs/nope/resume").status_code == 404

    def test_the_original_result_is_left_alone(self, client, store, stub_tool):
        """Resume produces a separate result and never edits the original.

        Merging is not attempted: inbound counts and orphan flags are properties
        of the whole graph, and a checkpoint holds URLs only.
        """
        stub_tool.result = StubResult(ok=True, data={"base_url": SAFE_URL})
        original = run_job(client, store)
        self._checkpoint(store, original, [f"{SAFE_URL}a"])

        before = store.get(original)
        resumed = client.post(f"{API_PREFIX}/jobs/{original}/resume").json()["id"]
        after = store.get(original)

        assert resumed != original
        # The whole record, not one field: resume must not touch the original at
        # all, and asserting on a single attribute would miss a status or
        # telemetry rewrite.
        assert after == before


class TestCancel:
    """Abandoning a job and reclaiming its slot.

    The failure this exists for: two crawls wedged in network I/O held two of
    three slots for sixteen hours, so every new crawl was refused by a server
    that was doing nothing useful.
    """

    def test_cancelling_frees_the_slot(self, store):
        app = create_app(
            store=store,
            url_policy=UrlSafetyPolicy(resolver=lambda host: [PUBLIC_IP]),
            max_concurrent_jobs=1,
        )
        state = app.state.api
        record = store.create("tool", {"base_url": "https://e.com/"}, label="stuck")

        # Marked running *inside* the client context. Startup recovers orphaned
        # jobs — anything left RUNNING by a dead process is failed on boot — so
        # a job put in that state beforehand is already terminal by the time the
        # request arrives, and the cancel would 409 for the wrong reason.
        with TestClient(app) as client:
            store.mark_running(record.id)
            state.try_reserve(record.id)
            response = client.post(f"{API_PREFIX}/jobs/{record.id}/cancel")

        assert response.status_code == 200
        assert state.active_count == 0
        assert store.get(record.id).status is JobStatus.FAILED

    def test_the_reason_says_the_thread_may_survive(self, store):
        """The button cannot kill a worker thread, and must not imply it can.

        `asyncio.to_thread` cannot be interrupted from outside, so the crawl
        keeps fetching until it finishes or the process exits. A cancelled job
        that read like a clean stop would be a lie told in the one place an
        operator goes when they already distrust the system.
        """
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = store.create("tool", {"base_url": "https://e.com/"}, label="stuck")

        with TestClient(app) as client:
            store.mark_running(record.id)
            client.post(f"{API_PREFIX}/jobs/{record.id}/cancel")

        assert "may still be running" in (store.get(record.id).error or "")

    def test_a_finished_job_cannot_be_cancelled(self, store):
        """409 rather than rewriting history. A finished crawl is the evidence."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = store.create("tool", {"base_url": "https://e.com/"}, label="done")
        store.finish(record.id, {"pages": []})

        with TestClient(app) as client:
            response = client.post(f"{API_PREFIX}/jobs/{record.id}/cancel")

        assert response.status_code == 409
        assert store.get(record.id).status is JobStatus.SUCCEEDED

    def test_cancelling_an_unknown_job_is_404(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            assert client.post(f"{API_PREFIX}/jobs/nope/cancel").status_code == 404


class TestHomepageSidecar:
    """The one page kept so a menu can be re-parsed later."""

    def test_a_homepage_round_trips(self, store):
        record = store.create("tool", {"base_url": "https://e.com/"})
        store.write_homepage(record.id, "<header><nav>menu</nav></header>")
        assert store.read_homepage(record.id) == "<header><nav>menu</nav></header>"

    def test_an_absent_homepage_is_none_not_an_error(self, store):
        """Every crawl older than the sidecar answers this way, and must."""
        record = store.create("tool", {"base_url": "https://e.com/"})
        assert store.read_homepage(record.id) is None

    def test_an_oversized_homepage_is_dropped_not_truncated(self, store):
        """Half a document parses into a menu that is quietly missing its tail.

        A wrong tree is worse than no tree, so the body is refused outright and
        the reparse falls back to the stored menu.
        """
        record = store.create("tool", {"base_url": "https://e.com/"})
        store.write_homepage(record.id, "x" * (MAX_HOMEPAGE_BYTES + 1))
        assert store.read_homepage(record.id) is None

    def test_writing_for_an_unknown_job_does_not_raise(self, store):
        """Losing a re-parsing aid must never fail the crawl that produced it."""
        store.write_homepage("nope", "<html></html>")


class TestReparseEndpoint:
    def _finished(self, store: DiskJobStore, client: TestClient) -> JobRecord:
        """A job carrying a real, contract-valid result."""
        record = store.create(
            server_module.TOOL_NAME, {"base_url": "https://e.com/"}, label="e.com"
        )
        store.finish(
            record.id,
            PageClassificationOutput(
                base_url="https://e.com/",
                site_profile=SiteProfile(),
                weight_profile=WeightProfileReport.for_site(SiteProfile()),
                discovery=DiscoveryReport(base_url="https://e.com/"),
                summary=CrawlSummary(),
                pages=(),
            ).model_dump(mode="json"),
        )
        return record

    def test_reparsing_creates_a_new_job_and_keeps_the_original(self, store):
        """The original is the evidence of what the site was when crawled."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store, client)
            response = client.post(f"{API_PREFIX}/jobs/{source.id}/reparse")

        assert response.status_code == 200
        body = response.json()
        assert body["id"] != source.id
        assert body["status"] == JobStatus.SUCCEEDED.value
        assert "(reparsed)" in body["label"]
        assert store.get(source.id).status is JobStatus.SUCCEEDED

    def test_the_menu_is_re_parsed_when_a_homepage_was_kept(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store, client)
            store.write_homepage(
                source.id,
                "<header><nav><ul><li><a href='https://e.com/docs/'>Docs</a></li></ul></nav></header>",
            )
            body = client.post(f"{API_PREFIX}/jobs/{source.id}/reparse").json()

        result = store.read_result(body["id"])
        assert [r["label"] for r in result["navigation"]["roots"]] == ["Docs"]
        # Carried forward, or the chain breaks after one hop.
        assert store.read_homepage(body["id"]) is not None

    def test_reparsing_without_a_homepage_keeps_the_stored_menu(self, store):
        """The normal case for every crawl older than the sidecar."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store, client)
            body = client.post(f"{API_PREFIX}/jobs/{source.id}/reparse").json()

        assert store.read_result(body["id"])["navigation"]["roots"] == []

    def test_reparsing_a_job_with_no_result_is_404(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            record = store.create(server_module.TOOL_NAME, {"base_url": "https://e.com/"})
            assert client.post(f"{API_PREFIX}/jobs/{record.id}/reparse").status_code == 404


class TestScreamingFrogEndpoint:
    """The optional reconciliation pass.

    Optional is the load-bearing word: no other endpoint calls this, and a crawl
    that never sees an export must behave exactly as it always has. The tests
    that matter most here are the ones asserting what it does *not* do.
    """

    HEADER = (
        "Address,Status Code,Content Type,Indexability,Indexability Status,"
        "Redirect URL,Crawl Depth,Unique Inlinks"
    )
    PATH = "reconcile/screaming-frog"

    def _finished(self, store: DiskJobStore) -> JobRecord:
        """A one-page crawl to reconcile against."""
        page = FullPageIntelligenceProfile(
            url="https://e.com/a/",
            canonical_url="https://e.com/a/",
            normalized_path="https://e.com/a/",
            hierarchy_level=HierarchyLevel.L3_LEAF_PAGE,
            primary_page_type=PrimaryPageType.BLOG_ARTICLE,
            depth_from_l0=1,
            search_intent=SearchIntent.INFORMATIONAL,
            signals_evaluated=(
                SignalScore(
                    source=SignalSource.SITEMAP_INDEX,
                    suggested_level=HierarchyLevel.L3_LEAF_PAGE,
                    suggested_page_type=PrimaryPageType.BLOG_ARTICLE,
                    confidence=0.9,
                ),
            ),
            final_confidence_score=0.9,
            consensus_method=ConsensusMethod.LAYER1_STRUCTURAL,
        )
        record = store.create(
            server_module.TOOL_NAME, {"base_url": "https://e.com/"}, label="e.com"
        )
        store.finish(
            record.id,
            PageClassificationOutput(
                base_url="https://e.com/",
                site_profile=SiteProfile(),
                weight_profile=WeightProfileReport.for_site(SiteProfile()),
                discovery=DiscoveryReport(base_url="https://e.com/"),
                summary=CrawlSummary(pages_classified=1),
                pages=(page,),
            ).model_dump(mode="json"),
        )
        return record

    def _post(self, client: TestClient, job_id: str, *lines: str) -> httpx.Response:
        return client.post(
            f"{API_PREFIX}/jobs/{job_id}/{self.PATH}",
            content="\n".join((self.HEADER, *lines)).encode("utf-8"),
            headers={"Content-Type": "text/csv"},
        )

    def test_a_missed_page_is_merged_into_a_new_job(self, store):
        """The original stays put: the comparison is the point of running this."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store)
            response = self._post(
                client, source.id, '"https://e.com/b/",200,text/html,Indexable,,,3,5'
            )

        assert response.status_code == 200
        body = response.json()
        assert body["merged"] == 1
        assert body["missed_pages"] == 1
        assert body["job_id"] != source.id
        assert body["source_job_id"] == source.id
        # The source result is untouched.
        assert len(store.read_result(source.id)["pages"]) == 1
        assert len(store.read_result(body["job_id"])["pages"]) == 2

    def test_nothing_to_merge_creates_no_job(self, store):
        """A report-only run must not litter `.jobs/` with duplicates."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store)
            before = len(store.list_jobs())
            response = self._post(
                client, source.id, '"https://e.com/a/",200,text/html,Indexable,,,3,5'
            )
            after = len(store.list_jobs())

        body = response.json()
        assert body["merged"] == 0
        assert body["job_id"] == source.id
        assert after == before

    def test_noise_is_reported_but_never_merged(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store)
            response = self._post(
                client,
                source.id,
                '"https://e.com/gone/",404,text/html,Non-Indexable,,,4,0',
                '"https://e.com/hero.jpg",200,image/jpeg,Indexable,,,2,9',
            )

        body = response.json()
        assert body["merged"] == 0
        assert body["frog_reasons"]["CLIENT_ERROR"] == 1
        assert body["frog_reasons"]["MEDIA_URL"] == 1

    def test_an_empty_body_is_rejected(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store)
            response = client.post(
                f"{API_PREFIX}/jobs/{source.id}/{self.PATH}",
                content=b"",
                headers={"Content-Type": "text/csv"},
            )
        assert response.status_code == 400

    def test_an_unknown_job_is_404(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            response = self._post(
                client, "nope", '"https://e.com/b/",200,text/html,Indexable,,,3,5'
            )
        assert response.status_code == 404

    def test_a_byte_order_mark_does_not_blind_the_parser(self, store):
        """Screaming Frog writes a BOM; without `utf-8-sig` the export reads as empty."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            source = self._finished(store)
            payload = "\n".join((self.HEADER, '"https://e.com/b/",200,text/html,Indexable,,,3,5'))
            response = client.post(
                f"{API_PREFIX}/jobs/{source.id}/{self.PATH}",
                content=b"\xef\xbb\xbf" + payload.encode("utf-8"),
                headers={"Content-Type": "text/csv"},
            )
        assert response.json()["merged"] == 1


class TestReconciliationIsKept:
    """A cross-check must survive closing the dialog.

    It costs an export somebody produced by hand in another tool, and it lived
    only in one React component's state: navigating away discarded it, and
    getting it back meant re-exporting and re-uploading several megabytes.
    """

    def test_a_saved_cross_check_round_trips(self, store):
        record = store.create("tool", {"base_url": "https://e.com/"})
        store.write_reconciliation(record.id, {"summary": {"in_both": 7}, "missed_pages": ["a"]})
        assert store.read_reconciliation(record.id) == {
            "summary": {"in_both": 7},
            "missed_pages": ["a"],
        }

    def test_a_job_nobody_cross_checked_reads_as_none(self, store):
        record = store.create("tool", {"base_url": "https://e.com/"})
        assert store.read_reconciliation(record.id) is None

    def test_writing_for_an_unknown_job_does_not_raise(self, store):
        """Losing the ability to re-read one must not fail the request."""
        store.write_reconciliation("nope", {"summary": {}})

    def test_the_endpoint_returns_the_saved_lists_not_just_counts(self, store):
        """The counts are the headline; the addresses are the work."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = store.create(server_module.TOOL_NAME, {"base_url": "https://e.com/"})
        store.write_reconciliation(
            record.id,
            {"summary": {"in_both": 2}, "missed_pages": ["https://e.com/x"], "orphans": []},
        )
        with TestClient(app) as client:
            body = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation").json()
        assert body["missed_pages"] == ["https://e.com/x"]

    def test_an_uncross_checked_job_is_404(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = store.create(server_module.TOOL_NAME, {"base_url": "https://e.com/"})
        with TestClient(app) as client:
            response = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation")
        assert response.status_code == 404

    def test_a_sidecar_is_not_mistaken_for_a_job(self, store):
        """`list_jobs` reads `.jobs/` by filename and has been fooled before.

        A checkpoint sidecar was once read as a job record, which meant every
        request parsed 102 MB of JSON.
        """
        record = store.create("tool", {"base_url": "https://e.com/"})
        store.write_reconciliation(record.id, {"summary": {}})
        assert [job.id for job in store.list_jobs()] == [record.id]


class TestReconciliationDownload:
    """The artefact a person keeps, as opposed to the JSON the panel redraws from."""

    def _saved(self, store: DiskJobStore) -> JobRecord:
        record = store.create(server_module.TOOL_NAME, {"base_url": "https://e.com/"})
        store.write_reconciliation(
            record.id,
            {
                "summary": {"base_url": "https://e.com/", "in_both": 2, "missed_pages": 1},
                "created_at": "2026-08-20T09:00:00+00:00",
                "frog_only": [{"url": "https://e.com/x", "reason": "MISSED_PAGE"}],
                "engine_only": [{"url": "https://e.com/y", "reason": "SITEMAP_ORPHAN"}],
            },
        )
        return record

    def test_both_sides_land_in_one_table(self, store):
        """The question is per URL: which crawler saw this, and why not the other?

        Two files would make the reader do a join to answer it.
        """
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            body = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.csv").text

        assert "https://e.com/x,screaming_frog_only,MISSED_PAGE" in body
        assert "https://e.com/y,rankuno_only,SITEMAP_ORPHAN" in body

    def test_the_counts_ride_along(self, store):
        """Without them a reader treats the gap lists as the whole site."""
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            body = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.csv").text
        assert "2,summary,in_both," in body

    def test_reasons_are_glossed_for_a_reader_who_is_not_an_engineer(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            body = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.csv").text
        assert "no internal link reaches it" in body

    def test_the_workbook_splits_the_lists_into_sheets(self, store):
        """One sheet per question, and the actionable lists come first.

        The flat CSV was the right call when the argument was "two files force a
        join". Sheets in one workbook are not two files — nothing has to be
        joined — and a real gep.com cross-check is 17,640 rows, in which the 15
        pages the crawl actually missed sat below sixteen thousand differences
        that need no action.
        """
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            response = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.xlsx")

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument"
        )
        book = load_workbook(io.BytesIO(response.content))
        # One sheet per reason, the two findings first. Both sides land in the
        # same list because `FrogGapReason` and `EngineGapReason` share no
        # member, so a reason already names its side.
        assert book.sheetnames == ["Summary", "Missed pages", "Orphans"]
        assert [cell.value for cell in book["Missed pages"][2]] == ["https://e.com/x"]
        assert [cell.value for cell in book["Orphans"][2]] == ["https://e.com/y"]

    def test_a_reason_sheet_carries_urls_and_nothing_else(self, store):
        """The reason is the sheet; repeating it down the rows is noise.

        The flat version put `SITEMAP_ORPHAN` and its gloss in every one of 801
        rows, and `MEDIA_URL` in 16,162 — the same two values, over and over,
        occupying the two columns beside the only one that varies.
        """
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            response = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.xlsx")
        book = load_workbook(io.BytesIO(response.content))
        assert [cell.value for cell in book["Orphans"][1]] == ["url"]
        assert book["Orphans"].max_column == 1

    def test_the_summary_is_a_contents_page(self, store):
        """Where each reason's meaning lives now that it is not on every row.

        With a dozen sheets the reader has to know which to open, and the tab
        strip says `Media files` without saying that it is 16,162 URLs the
        engine refused on purpose.
        """
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            response = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.xlsx")
        rows = list(
            load_workbook(io.BytesIO(response.content))["Summary"].iter_rows(values_only=True)
        )
        contents = [row for row in rows if row and row[0] == "Missed pages"]
        assert contents, "the contents page names every sheet"
        assert contents[0][1] == "Screaming Frog"
        assert contents[0][2] == 1
        assert "did not reach it" in str(contents[0][3])

    def test_every_sheet_keeps_its_header_in_view(self, store):
        """A 16,000-row sheet whose header scrolls away is unreadable.

        `freeze_panes` is silently dropped by a write-only worksheet if it is set
        after the first row is appended — openpyxl accepts the assignment and
        discards it. The first version of this endpoint did exactly that.
        """
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            response = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.xlsx")

        book = load_workbook(io.BytesIO(response.content))
        assert all(sheet.freeze_panes == "A2" for sheet in book.worksheets)
        # And a URL column wide enough to read a URL in.
        assert book["Orphans"].column_dimensions["A"].width == 60

    def test_the_workbook_needs_a_saved_cross_check(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        with TestClient(app) as client:
            assert client.get(f"{API_PREFIX}/jobs/nope/reconciliation.xlsx").status_code == 404

    def test_it_downloads_as_a_named_file(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = self._saved(store)
        with TestClient(app) as client:
            response = client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.csv")
        assert response.headers["content-disposition"].startswith("attachment;")
        assert "2026-08-20.csv" in response.headers["content-disposition"]

    def test_a_job_never_cross_checked_is_404(self, store):
        app = create_app(store=store, url_policy=UrlSafetyPolicy(resolver=lambda h: [PUBLIC_IP]))
        record = store.create(server_module.TOOL_NAME, {"base_url": "https://e.com/"})
        with TestClient(app) as client:
            assert (
                client.get(f"{API_PREFIX}/jobs/{record.id}/reconciliation.csv").status_code == 404
            )
