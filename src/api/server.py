"""FastAPI application serving crawl jobs to the React UI.

Why jobs and not a synchronous endpoint
---------------------------------------
A 20k-page crawl takes minutes. It cannot run inside a request: the browser, and
every proxy between it and here, will time out long before it finishes. So
`POST /jobs` accepts the work, returns `202` with an id, and the client polls.

Why a worker thread and not `BackgroundTasks`
---------------------------------------------
`PageClassificationTool.execute()` is synchronous and calls `asyncio.run()`
internally. Running it in FastAPI's `BackgroundTasks` would execute it *on the
event loop*, blocking every other request for the whole crawl — including the
`GET /jobs/{id}` polls the UI needs to show progress. It is dispatched to a
worker thread with `asyncio.to_thread` instead, which is also what lets the
concurrency cap mean anything.

What this layer does not do
---------------------------
No safety control is implemented here. SSRF validation, robots compliance,
per-host throttling, guardrails and audit logging all live in the tool and its
fetcher. The one apparent exception is the URL check on `POST /jobs`, which is
not a second guard but an early one: the same `UrlSafetyPolicy` the crawl will
use anyway, run at admission so a bad URL is a `400` the operator sees
immediately rather than a job that fails a moment later.

Binding
-------
`serve()` binds `127.0.0.1` deliberately. This server has no authentication, and
it will fetch arbitrary URLs on request — on a routable interface that is an
open proxy. Local-only is the security boundary (ADR 0004).
"""

from __future__ import annotations

import asyncio
import csv
import io
import threading
import time
from collections.abc import AsyncGenerator, Mapping, Sequence
from contextlib import asynccontextmanager
from datetime import UTC, datetime
from pathlib import Path
from types import MappingProxyType
from uuid import uuid4

from fastapi import FastAPI, HTTPException, Request, Response, status
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import Field, ValidationError

from src.core.errors import UnsafeUrlError
from src.core.logger import get_logger
from src.core.schemas import StrictModel
from src.core.state_store import (
    MAX_RECENT_ITEMS,
    DiskJobStore,
    JobNotFoundError,
    JobRecord,
    JobStatus,
    JobStore,
    JobTelemetry,
)
from src.core.url_safety import UrlSafetyPolicy
from src.modules.seo.page_classifier.discovery import DiscoveryReport, SiteGraph
from src.modules.seo.page_classifier.schemas import (
    ConsensusMethod,
    FullPageIntelligenceProfile,
    HierarchyLevel,
    PrimaryPageType,
    SearchIntent,
    SignalScore,
    SignalSource,
)
from src.modules.seo.page_classifier.screaming_frog_merge import (
    merge_reconciled_urls,
)
from src.modules.seo.page_classifier.tool import (
    CrawlSummary,
    PageClassificationInput,
    PageClassificationOutput,
    PageClassificationTool,
    reparse_placement,
)
from src.modules.seo.page_classifier.url_rules import safe_split, site_host
from src.modules.seo.page_classifier.weights import SiteProfile, WeightProfileReport
from src.modules.seo.performance.aggregator import (
    PageMetricSet,
    UnmatchedGroup,
    merge_page_metrics,
    rollup_of,
    section_path_of,
    unmatched_groups,
)
from src.modules.seo.performance.gsc_export import load_gsc_export
from src.modules.seo.performance.opportunity_scorer import (
    OpportunityReport,
    score_opportunities,
)
from src.modules.seo.performance.schemas import (
    GscPageMetrics,
    MatchFailure,
    PerformanceRollup,
)
from src.modules.seo.performance.url_identity import UrlResolutionIndex

__all__ = ["ApiState", "CrawlCheckpointer", "TelemetryRecorder", "create_app", "serve"]

_logger = get_logger("api.server")

API_PREFIX = "/api/v1"

API_VERSION = "0.1.0"
"""Version of the HTTP contract, not of the engine.

Separate on purpose: the engine's classification can improve without the request
and response shapes changing, and a client pins to the shapes.
"""

DEFAULT_MAX_CONCURRENT_JOBS = 3
"""Simultaneous crawls this process will run.

Not about local CPU — the fetcher's token bucket already bounds politeness per
host. It bounds *memory*: each in-flight crawl holds its whole graph, including
page HTML, until it finishes. Three 20k-page crawls at once is already several
gigabytes.
"""

DEFAULT_ALLOWED_ORIGINS = (
    "http://localhost:5173",
    "http://127.0.0.1:5173",
)
"""The Vite dev server, and nothing else.

Not `*`. A wildcard would let any page the operator happens to be visiting read
crawl results out of this server.
"""

TOOL_NAME = "seo.page_classifier"

TELEMETRY_FLUSH_SECONDS = 0.5
"""Minimum gap between telemetry writes.

Not a preference — a correctness bound. Each write rewrites the job record
through `os.replace` and an `fsync`. Flushing per page on a 20,000-page crawl
would spend more wall-clock in the filesystem than on the network, so the
telemetry would slow the thing it is measuring.
"""

TELEMETRY_WARMUP_SECONDS = 3.0
"""Elapsed time before an ETA is offered at all.

A rate computed over the first fraction of a second is noise, and an ETA derived
from it swings between seconds and hours. Showing nothing is better than showing
a number that is about to change by two orders of magnitude.
"""

RATE_SMOOTHING = 0.3
"""EMA weight for the newest rate sample, 0–1.

Lower is smoother. At 0.3 a single stalled request moves the estimate a little;
an instantaneous rate would make it jump on every retry.
"""


CHECKPOINT_INTERVAL_S = 10.0
"""Minimum gap between checkpoint writes."""

CHECKPOINT_EVERY_PAGES = 100
"""Pages between checkpoints, whichever boundary arrives first.

Two triggers because either alone fails at one end of the range. Time alone
under-saves a fast crawl — Turbo at 25 pages/sec puts 250 pages between writes.
Page count alone under-saves a slow one, where 100 pages can be several minutes
of work exposed to a power cut.
"""


class CrawlCheckpointer:
    """Saves the discovered URL set so an interruption does not lose the crawl.

    Deliberately **not** the classified output. Re-serialising every profile on
    each write means megabytes through `fsync` hundreds of times per crawl —
    measurably more expensive than the crawling. URLs are what cannot be
    recovered without going back to the network; classification is CPU-only and
    can be redone.

    Called from crawler threads, so mutations are under the lock.
    """

    def __init__(self, store: JobStore, job_id: str, base_url: str) -> None:
        """Build a checkpointer for one job."""
        self._store = store
        self._job_id = job_id
        self._base_url = base_url
        self._lock = threading.Lock()
        self._last_write = 0.0
        self._last_count = 0

    def __call__(self, graph: SiteGraph) -> None:
        """Save the graph if enough time or enough pages have passed."""
        now = time.monotonic()
        with self._lock:
            count = len(graph)
            due = (
                now - self._last_write >= CHECKPOINT_INTERVAL_S
                or count - self._last_count >= CHECKPOINT_EVERY_PAGES
            )
            if not due or count == 0:
                return
            self._last_write = now
            self._last_count = count
            # Materialised inside the throttle, never outside it: at 20,000
            # nodes building these tuples is the expensive part of a checkpoint.
            urls = graph.all_urls()
            unfetched = graph.unfetched_urls()

        try:
            self._store.write_checkpoint(
                self._job_id,
                {
                    "base_url": self._base_url,
                    "urls": list(urls),
                    # What a resumed crawl would still have to fetch. Recorded
                    # here because it cannot be recovered afterwards: the result
                    # holds a row for every *discovered* URL, fetched or not, so
                    # nothing downstream can tell the two apart.
                    "unfetched": list(unfetched),
                    "saved_at_count": count,
                },
            )
        except JobNotFoundError:
            _logger.debug("checkpoint_job_missing", extra={"job_id": self._job_id})


class TelemetryRecorder:
    """Turns raw progress callbacks into a throttled, smoothed snapshot.

    Lives in the API layer, not in the crawler: throughput smoothing and write
    throttling are presentation concerns for a polling client, and the engine
    should not know that anyone is watching.

    Called from crawler threads, so every mutation is under the lock.
    """

    def __init__(self, store: JobStore, job_id: str, ceiling: int) -> None:
        """Build a recorder for one job.

        Args:
            store: Where snapshots are persisted.
            job_id: The job being reported on.
            ceiling: `max_pages`, used only to cap the denominator — never as
                the denominator itself.
        """
        self._store = store
        self._job_id = job_id
        self._ceiling = ceiling
        self._lock = threading.Lock()
        self._started = time.monotonic()
        self._last_flush = 0.0
        self._last_completed = 0
        self._last_sample = self._started
        self._rate = 0.0

    def __call__(self, completed: int, discovered: int, recent: tuple[str, ...]) -> None:
        """Record progress. Cheap, non-blocking, and never raises."""
        now = time.monotonic()
        with self._lock:
            self._update_rate(completed, now)
            if now - self._last_flush < TELEMETRY_FLUSH_SECONDS and completed > 0:
                return
            self._last_flush = now
            snapshot = self._snapshot(completed, discovered, recent, now)

        try:
            self._store.update_telemetry(self._job_id, snapshot)
        except JobNotFoundError:
            # The job was deleted mid-crawl. Losing telemetry is not a reason to
            # interrupt work that is still producing a result.
            _logger.debug("telemetry_job_missing", extra={"job_id": self._job_id})

    def _update_rate(self, completed: int, now: float) -> None:
        elapsed = now - self._last_sample
        if elapsed < 0.05:
            return
        sample = max(0, completed - self._last_completed) / elapsed
        # Seeded rather than blended on the first sample: blending against a
        # zero start would halve the very first estimate.
        self._rate = (
            sample
            if self._rate == 0.0
            else (RATE_SMOOTHING * sample + (1 - RATE_SMOOTHING) * self._rate)
        )
        self._last_completed = completed
        self._last_sample = now

    def _snapshot(
        self, completed: int, discovered: int, recent: tuple[str, ...], now: float
    ) -> JobTelemetry:
        # Capped at the ceiling because the crawl stops there; measured against
        # what was discovered because that is the real total. Using the ceiling
        # as the denominator leaves a 300-page site reading 1.5% forever.
        total = min(discovered, self._ceiling) if discovered else 0
        remaining = max(0, total - completed)

        eta: float | None = None
        if now - self._started >= TELEMETRY_WARMUP_SECONDS and self._rate > 0 and total > 0:
            eta = remaining / self._rate

        return JobTelemetry(
            completed=completed,
            discovered=total,
            rate_per_sec=round(self._rate, 3),
            eta_seconds=None if eta is None else round(eta, 1),
            recent_items=recent[-MAX_RECENT_ITEMS:],
            updated_at=datetime.now(UTC),
        )


class HealthView(StrictModel):
    """Liveness response."""

    status: str = "ok"
    active_jobs: int = 0
    max_concurrent_jobs: int = DEFAULT_MAX_CONCURRENT_JOBS


GAP_MEANINGS: Mapping[str, str] = MappingProxyType(
    {
        "REDIRECT": "Redirect source. Not a page; the destination is in both crawls.",
        "OFF_SITE": "A different host. This engine is same-site by design.",
        "CLIENT_ERROR": "4xx or 5xx when crawled. Not a page.",
        "MEDIA_URL": "Image, stylesheet or script. Refused deliberately.",
        "SPIDER_TRAP": "Refused by this engine's trap rules.",
        "NON_INDEXABLE": "Live but canonicalised elsewhere or noindex.",
        "MISSED_PAGE": "Live, indexable, in scope - and this engine did not reach it.",
        "SITEMAP_ORPHAN": (
            "Published but no internal link reaches it. A link crawler cannot see these."
        ),
        "REPEATED_SUFFIX_TRAP": "One page at many fabricated addresses, from a relative href.",
        "MALFORMED_MARKUP": "Built from broken HTML on the site. Never a URL.",
        "QUERY_VARIANT": "The same path with a query string the other crawler collapsed.",
    }
)
"""Plain-language gloss for each gap reason, for the downloadable report.

The enum names are precise and mean nothing to the client who receives the
spreadsheet. Kept here rather than in the reconciler because this is presentation
- the module that decides the reasons should not also own how they read.
"""


MAX_PERFORMANCE_UPLOAD = 32 * 1024 * 1024
"""Largest Search Console upload accepted, before decompression.

The UI export caps at 1,000 rows and the API at 50,000, so a real export is
kilobytes to a few megabytes. This bounds what an accident or a hostile client
can push into memory; `gsc_export.MAX_UNPACKED_BYTES` bounds what it becomes
once unpacked, which is the separate question a compressed archive raises.
"""


UNMATCHED_MEANINGS: Mapping[str, str] = MappingProxyType(
    {
        "other_subdomain": (
            "A subdomain of this site that the crawl never covered. Worth a look "
            "on its own: an uncrawled property, a staging host that escaped, or a "
            "compromised one."
        ),
        "off_site": "A different domain entirely — another property.",
        "not_crawled": "On this site, but this crawl never reached it.",
        "ambiguous": "Several crawled pages claim this address; attributing it would be a guess.",
        "unparseable": "Not a URL. A malformed export cell.",
    }
)
"""Plain-language gloss for each unmatched reason, for the download.

Same reasoning as `GAP_MEANINGS`: the enum names are precise and mean nothing to
whoever receives the spreadsheet.
"""


def _matched_rows(index: UrlResolutionIndex, metrics: PageMetricSet) -> list[dict[str, object]]:
    """Every crawled page an export row reached, with the crawl's own columns.

    The working dataset, and the reason it is stored rather than derived: the
    summary records *which* pages matched but not what they earned, so a
    download built from it could list addresses and nothing else.

    Carries the section trail, page type and inbound-link count alongside the
    Search Console figures, because the join is the whole point — clicks next to
    where the page sits in the navigation is a thing a spreadsheet cannot
    produce on its own.
    """
    profiles = {page.url: page for page in index.pages}
    # Sorted on the counter rather than on the rendered dict: reading `clicks`
    # back out of a `dict[str, object]` is an unchecked cast, and the value has
    # already been turned into a display string one field along.
    rows: list[tuple[int, dict[str, object]]] = []
    for page_url, held in metrics.pages.items():
        gsc = held.gsc
        if gsc is None:
            continue
        profile = profiles.get(page_url)
        rows.append(
            (
                gsc.clicks,
                {
                    "url": page_url,
                    "section": " > ".join(section_path_of(profile)) if profile else "",
                    "clicks": gsc.clicks,
                    "impressions": gsc.impressions,
                    "ctr": round(gsc.ctr, 6),
                    "position": round(gsc.position, 2) if gsc.impressions else "",
                    "page_type": profile.primary_page_type.value if profile else "",
                    "hierarchy_level": profile.hierarchy_level.value if profile else "",
                    "inbound_internal_links": (
                        profile.inbound_internal_links_count if profile else ""
                    ),
                },
            )
        )
    return [row for _, row in sorted(rows, key=lambda pair: -pair[0])]


def _unmatched_rows(metrics: PageMetricSet) -> list[tuple[GscPageMetrics, str]]:
    """Each unresolved Search Console row paired with why it did not resolve."""
    reasons = {f.google_url: f.reason.value for f in metrics.gsc_resolution.failures}
    return [
        (row, reasons.get(row.url, MatchFailure.NOT_CRAWLED.value))
        for row in metrics.unresolved_gsc
    ]


Sheet = tuple[str, Sequence[str], Sequence[Sequence[object]]]
"""One worksheet: its name, its header row, and its data rows."""


def _workbook_response(sheets: Sequence[Sheet], filename: str) -> Response:
    """Render sheets as an `.xlsx` a person can actually work in.

    A CSV is one table, and the cross-check is five: a summary, two lists an
    analyst acts on, and two long lists they mostly scroll past. Flattened into
    one sheet with a `found_by` column — which is what cycle 0028 chose, for the
    good reason that splitting into separate *files* forces a join — a real
    report came out at 17,640 rows with the 14 that matter buried near the
    bottom. Sheets in one workbook are not separate files; nothing has to be
    joined, and each list is a click away.

    `write_only` because these run to tens of thousands of rows: the normal mode
    builds a cell object per value and holds the lot.

    Sheets are ordered smallest and most actionable first. That order is the
    argument of the report — the 14 pages the crawl missed are the finding, and
    16,465 rows of explained differences are the evidence behind it.
    """
    book = Workbook(write_only=True)
    for title, headers, rows in sheets:
        # Excel refuses a sheet name over 31 characters or holding []:*?/\ ,
        # and openpyxl raises rather than truncating.
        sheet = book.create_sheet(title[:31])
        # Both of these must be set *before* the first `append`. A write-only
        # sheet accepts them afterwards and silently discards them — verified,
        # because the first version of this did exactly that and shipped a
        # 16,000-row sheet whose header scrolled away, which is the one thing
        # this endpoint exists to fix.
        sheet.freeze_panes = "A2"
        for index, width in enumerate(_column_widths(headers), start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.append(list(headers))
        for row in rows:
            sheet.append(list(row))

    buffer = io.BytesIO()
    book.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _column_widths(headers: Sequence[str]) -> list[int]:
    """Column widths by what the header says the column holds.

    Guessed from the name rather than measured from the data, which would mean
    reading every row twice. A URL column at the default width shows about
    eight characters, which is the difference between a usable report and one
    the reader has to resize before they can start.
    """
    wide = {"url", "meaning", "reason", "section", "base_url"}
    return [
        60 if header.lower() in {"url", "meaning"} else 22 if header.lower() in wide else 14
        for header in headers
    ]


def _csv_response(body: str, filename: str) -> Response:
    """A CSV a spreadsheet will open with its accents intact.

    The byte-order mark is the whole point. Excel reads a `.csv` in the system
    codepage unless a BOM says otherwise, so UTF-8 without one arrives as
    mojibake — the first real report handed back `EspaÃ±ol` and `PortuguÃ©s`
    for gep.com's locale sections, in a file meant for a client. The
    `charset=utf-8` in the media type does not reach Excel; only the BOM does.
    """
    return Response(
        content="﻿" + body,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class PerformanceSummary(StrictModel):
    """What a Search Console upload produced against one crawl.

    The resolution figures come first deliberately. Every number below them is
    derived from a join between Google's URLs and ours, and a reader who sees
    the section totals without knowing that a third of the export failed to
    resolve is reading a confident understatement.
    """

    job_id: str
    base_url: str

    source_name: str = ""
    """The archive entry or worksheet the rows were read from. Reported because
    the choice is made by inspecting content rather than names, and a silent
    choice is one nobody can check."""

    rows: int = 0
    """Rows read from the export."""

    skipped_rows: int = 0
    """Rows in that table with no usable address."""

    matched: int = 0
    match_rate_pct: float = 0.0
    is_reliable: bool = False
    """Whether the match rate cleared the threshold. False means the totals
    below understate traffic by an unknown amount that is not evenly spread."""

    pages_with_data: int = 0
    """Crawled pages the export reached. Against `pages`, this is the coverage
    question the match rate cannot answer: a 1,000-row UI export against a
    12,000-page site resolves perfectly and still describes 8% of the site."""

    pages: int = 0

    unmatched: tuple[UnmatchedGroup, ...] = ()
    """Where the rows that reached no page went, by host and reason.

    Carried beside the match rate rather than behind a second request. A rate
    on its own asks to be taken on trust; these groups partition the unresolved
    rows exactly, so it can be checked instead."""

    rollup: PerformanceRollup = PerformanceRollup()
    opportunities: OpportunityReport = OpportunityReport()


SHEET_TITLES: Mapping[str, str] = MappingProxyType(
    {
        # The two findings.
        "MISSED_PAGE": "Missed pages",
        "SITEMAP_ORPHAN": "Orphans",
        # Everything else is a difference with an explanation.
        "CLIENT_ERROR": "4xx and 5xx",
        "REDIRECT": "Redirect sources",
        "NON_INDEXABLE": "Noindex or canonicalised",
        "OFF_SITE": "Off-site",
        "MEDIA_URL": "Media files",
        "SPIDER_TRAP": "Crawl traps",
        "QUERY_VARIANT": "Query variants",
        "REPEATED_SUFFIX_TRAP": "Loop URLs",
        "MALFORMED_MARKUP": "Malformed markup",
    }
)
"""Worksheet name per gap reason.

Plain language, because a tab strip is read at a glance and `REPEATED_SUFFIX_
TRAP` is not. Every name is well inside Excel's 31-character limit, which
`_workbook_response` truncates to rather than raising — a reason added later
with a long title would quietly lose its ending, so keep them short here.

The names need no side prefix: `FrogGapReason` and `EngineGapReason` share no
member, so a reason already says which crawler saw the URL. The contents page
states it anyway, for a reader who does not know that.
"""


class ReconciliationSummary(StrictModel):
    """What a Screaming Frog reconciliation found, and what it did about it.

    Returned instead of a bare `JobRecord` because the counts are the point: an
    operator uploads an export to learn the size of the gap, and half of them
    will not want the merged job at all. The new job's id is included so the UI
    can open it when they do.
    """

    job_id: str
    """The merged result, saved as a new job. Equal to `source_job_id` when
    nothing was merged — there is no new job in that case."""

    source_job_id: str
    base_url: str

    frog_rows: int = 0
    """Rows read from the export."""

    in_both: int = 0
    """URLs both crawlers found."""

    missed_pages: int = 0
    """Live, indexable, in-scope pages the engine never reached. Merged."""

    orphans: int = 0
    """Published pages no internal link reaches. Found only by the engine, and
    left where they are — their absence from the export *is* the finding."""

    merged: int = 0
    """Pages added to the tree. Zero is a normal outcome."""

    frog_reasons: Mapping[str, int] = Field(default_factory=dict)
    """Why each frog-only URL was not merged, by reason."""

    engine_reasons: Mapping[str, int] = Field(default_factory=dict)
    """Why each engine-only URL is absent from the export, by reason."""


class JobAccepted(StrictModel):
    """What `POST /jobs` returns: an id to poll, not a result."""

    id: str
    status: str
    label: str = ""


class ApiState:
    """Everything the endpoints need, built once per application.

    Held on `app.state` rather than in module globals so a test can construct an
    isolated app — its own job directory, its own resolver — without touching
    the process the next test runs in.
    """

    def __init__(
        self,
        store: JobStore,
        url_policy: UrlSafetyPolicy,
        max_concurrent_jobs: int = DEFAULT_MAX_CONCURRENT_JOBS,
    ) -> None:
        """Build the shared state.

        Args:
            store: Job persistence.
            url_policy: SSRF policy used at admission and by the crawl.
            max_concurrent_jobs: Simultaneous crawls before requests are refused.
        """
        self.store = store
        self.url_policy = url_policy
        self.max_concurrent_jobs = max_concurrent_jobs
        self._active: set[str] = set()
        self._lock = threading.Lock()
        # Strong references to in-flight tasks. `asyncio` only holds weak ones,
        # so without this the garbage collector may cancel a running crawl.
        self._tasks: set[asyncio.Task[None]] = set()

    @property
    def active_count(self) -> int:
        """Crawls running right now."""
        with self._lock:
            return len(self._active)

    def try_reserve(self, job_id: str) -> bool:
        """Claim a concurrency slot, or report that none is free.

        Checked and claimed under one lock. Testing capacity and then reserving
        in two steps would let two simultaneous requests both pass the check.
        """
        with self._lock:
            if len(self._active) >= self.max_concurrent_jobs:
                return False
            self._active.add(job_id)
            return True

    def release(self, job_id: str) -> None:
        """Give a concurrency slot back."""
        with self._lock:
            self._active.discard(job_id)

    def rekey(self, provisional: str, job_id: str) -> None:
        """Move a reservation from a provisional id onto the real one.

        Capacity has to be claimed before the store mints an id, or a refusal
        leaves a persisted job behind. This transfers the claim without ever
        dropping it, so the slot cannot be taken in between.
        """
        with self._lock:
            self._active.discard(provisional)
            self._active.add(job_id)

    def is_active(self, job_id: str) -> bool:
        """Whether this job currently holds a concurrency slot."""
        with self._lock:
            return job_id in self._active

    def track(self, task: asyncio.Task[None]) -> None:
        """Hold a strong reference to an in-flight task until it completes.

        `asyncio` keeps only weak references to tasks, so a crawl with no other
        referent can be garbage-collected mid-run — the job would simply stop,
        leaving no error and no result.
        """
        self._tasks.add(task)
        task.add_done_callback(self._tasks.discard)


def _run_job(state: ApiState, job_id: str, payload: PageClassificationInput) -> None:
    """Execute one crawl to completion. Runs on a worker thread.

    Never raises: this runs detached, so an exception escaping here would be
    logged by asyncio and leave the job `RUNNING` forever with nothing to move
    it. Every path ends in a terminal status.
    """
    store = state.store
    try:
        store.mark_running(job_id)
        result = PageClassificationTool(
            url_policy=state.url_policy,
            progress_sink=TelemetryRecorder(store, job_id, payload.resolved_max_pages),
            checkpoint_sink=CrawlCheckpointer(store, job_id, payload.base_url),
            # Kept so a later fix to the header-menu parser can be applied to
            # this result without re-crawling. One page; the menu is global.
            homepage_sink=lambda html: store.write_homepage(job_id, html),
        ).run(payload)

        if not result.ok or result.data is None:
            store.mark_failed(job_id, result.error or "the tool returned no data")
            return

        output = result.data
        if not isinstance(output, PageClassificationOutput):  # pragma: no cover - defensive
            store.mark_failed(job_id, f"unexpected output type {type(output).__name__}")
            return

        # `truncated` means the crawl hit its ceiling, so the graph is a subset
        # of the site. Recorded as PARTIAL so the UI can say so rather than
        # presenting an incomplete crawl as a finished one.
        store.finish(
            job_id,
            output.model_dump(mode="json"),
            partial=output.discovery.truncated,
        )
    except Exception as exc:  # noqa: BLE001 - a detached worker must not leak
        _logger.exception("job_crashed", extra={"job_id": job_id})
        store.mark_failed(job_id, f"{type(exc).__name__}: {exc}")


async def _dispatch(state: ApiState, job_id: str, payload: PageClassificationInput) -> None:
    """Run a job on a worker thread and always release its slot."""
    try:
        await asyncio.to_thread(_run_job, state, job_id, payload)
    finally:
        state.release(job_id)


def create_app(
    store: JobStore | None = None,
    url_policy: UrlSafetyPolicy | None = None,
    *,
    jobs_root: Path | str | None = None,
    max_concurrent_jobs: int = DEFAULT_MAX_CONCURRENT_JOBS,
    allowed_origins: tuple[str, ...] = DEFAULT_ALLOWED_ORIGINS,
) -> FastAPI:
    """Build the application.

    A factory, not a module-level singleton, so tests get an isolated instance
    and so importing this module has no side effects — a module-level app would
    create the jobs directory and run orphan recovery at import time, including
    during test collection.

    Args:
        store: Job persistence. Defaults to a `DiskJobStore` under `jobs_root`.
        url_policy: SSRF policy applied at admission and inherited by the crawl.
        jobs_root: Directory for the default store. Defaults to `.jobs/`.
        max_concurrent_jobs: Simultaneous crawls before `429`.
        allowed_origins: Exact CORS origins. Never a wildcard.

    Returns:
        The configured application.
    """
    resolved_store: JobStore = (
        store if store is not None else DiskJobStore(jobs_root or Path(".jobs"))
    )
    state = ApiState(
        store=resolved_store,
        url_policy=url_policy if url_policy is not None else UrlSafetyPolicy(),
        max_concurrent_jobs=max_concurrent_jobs,
    )

    @asynccontextmanager
    async def lifespan(_app: FastAPI) -> AsyncGenerator[None]:
        # A job left RUNNING by a killed process has no worker any more. Nothing
        # would ever move it, and a polling UI would wait forever.
        orphans = resolved_store.recover_orphans()
        if orphans:
            _logger.warning("recovered_orphaned_jobs", extra={"count": len(orphans)})
        yield

    app = FastAPI(
        title="Rankuno AI Engine",
        version=API_VERSION,
        summary="Local API for the Phase 1 Page Classification Engine.",
        lifespan=lifespan,
    )
    app.state.api = state

    app.add_middleware(
        CORSMiddleware,
        allow_origins=list(allowed_origins),
        allow_credentials=False,
        allow_methods=["GET", "POST"],
        allow_headers=["Content-Type"],
    )

    # The endpoints close over `state` rather than receiving it through
    # `Depends`. With `from __future__ import annotations` every annotation is a
    # string, and FastAPI resolves those against the *module* namespace — a
    # dependency alias defined inside this factory is invisible there, so the
    # parameter silently degrades into a required query parameter and every
    # request 422s. A closure has no such failure mode.

    @app.get(f"{API_PREFIX}/health", response_model=HealthView)
    def health() -> HealthView:
        """Liveness, plus how much crawl capacity is free."""
        return HealthView(
            active_jobs=state.active_count,
            max_concurrent_jobs=state.max_concurrent_jobs,
        )

    @app.post(
        f"{API_PREFIX}/jobs", response_model=JobAccepted, status_code=status.HTTP_202_ACCEPTED
    )
    async def create_job(payload: PageClassificationInput) -> JobAccepted:
        """Accept a crawl and return an id to poll.

        `202`, never `200`: nothing has been crawled when this returns.

        Raises:
            HTTPException: `400` if the URL fails SSRF validation, `429` if no
                concurrency slot is free.
        """
        return _start(payload, payload.base_url)

    @app.get(f"{API_PREFIX}/jobs", response_model=list[JobRecord])
    def list_jobs() -> list[JobRecord]:
        """Every job, newest first. Metadata only — never a result blob."""
        return state.store.list_jobs()

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}", response_model=JobRecord)
    def get_job(job_id: str) -> JobRecord:
        """One job's status.

        Raises:
            HTTPException: `404` if no such job exists.
        """
        try:
            return state.store.get(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"no job {job_id}") from exc

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/result")
    def get_result(job_id: str) -> Mapping[str, object]:
        """A finished job's `PageClassificationOutput`.

        Returned as the stored mapping rather than re-validated into the model:
        it was serialised *from* that model, and re-parsing 16 MB on every fetch
        buys nothing.

        Raises:
            HTTPException: `404` if unknown, `409` if the job has not finished.
        """
        try:
            record = state.store.get(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"no job {job_id}") from exc

        if not record.has_result:
            # 409, not 404: the job exists and may yet produce a result. A 404
            # would tell a polling client to give up.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=f"job {job_id} is {record.status.value} and has no result",
            )
        return state.store.read_result(job_id)

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/checkpoint")
    def get_checkpoint(job_id: str) -> Mapping[str, object]:
        """A renderable view of what a job saved before it ended.

        Shaped as a `PageClassificationOutput` so the client renders it through
        exactly the path it uses for a finished crawl — a second shape would
        mean a second set of components to keep in step.

        Every page comes back `UNKNOWN`, which is honest rather than lazy: a
        checkpoint stores URLs, not classifications. The structure is real; what
        each page *is* was never determined.

        Raises:
            HTTPException: `404` if the job or its checkpoint does not exist.
        """
        try:
            state.store.get(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"no job {job_id}") from exc

        checkpoint = state.store.read_checkpoint(job_id)
        if checkpoint is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND, detail=f"job {job_id} saved no partial work"
            )
        return _checkpoint_as_output(checkpoint).model_dump(mode="json")

    def _start(payload: PageClassificationInput, label: str) -> JobAccepted:
        """Admit a crawl, reserve a slot, and dispatch it.

        Shared by every endpoint that starts work so admission cannot drift
        between them. SSRF validation in particular must not become something
        one entry point does and another forgets — a retry re-runs a stored
        payload, and a stored payload is not automatically still safe: DNS moves,
        and a host that resolved publicly last week may resolve to a private
        address today.
        """
        try:
            state.url_policy.validate(payload.base_url)
        except UnsafeUrlError as exc:
            _logger.warning("job_rejected_unsafe_url", extra={"url": payload.base_url})
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

        # Capacity is claimed *before* the record exists. The other order
        # persisted a job for every refusal and immediately marked it failed, so
        # a user retrying against a full server manufactured a permanent FAILED
        # row per click — 16 of 99 records on one workstation were nothing but
        # refusals, indistinguishable in the list from crawls that really ran.
        #
        # The slot is reserved against a provisional id and re-keyed once the
        # record exists, because `try_reserve` needs something to hold and the
        # id is the store's to mint.
        pending = f"pending:{uuid4().hex}"
        if not state.try_reserve(pending):
            raise HTTPException(
                status.HTTP_429_TOO_MANY_REQUESTS,
                detail=f"at most {state.max_concurrent_jobs} crawls may run at once",
            )
        try:
            record = state.store.create(TOOL_NAME, payload.model_dump(mode="json"), label=label)
        except Exception:
            # The store failed, so there is no job and nothing will ever release
            # this slot. Leaking it would cost a permanent slot per failure.
            state.release(pending)
            raise
        state.rekey(pending, record.id)
        state.track(asyncio.create_task(_dispatch(state, record.id, payload)))
        return JobAccepted(id=record.id, status=record.status.value, label=record.label)

    def _stored_payload(job_id: str) -> PageClassificationInput:
        """Rebuild the input a job was started with.

        Raises:
            HTTPException: `404` if there is no such job, `409` if its stored
                request cannot be replayed.
        """
        try:
            record = state.store.get(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"no job {job_id}") from exc

        if not record.request:
            # Jobs created before the request was stored, and any record written
            # by a different tool. Refused rather than guessed at: inventing a
            # payload would crawl with settings the operator never chose.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=f"job {job_id} did not record the request it ran, so it cannot be re-run",
            )
        try:
            return PageClassificationInput.model_validate(record.request)
        except ValidationError as exc:
            # A record written before a schema change. Says so plainly instead
            # of surfacing a pydantic traceback to the operator.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=f"job {job_id} was run with settings this build no longer accepts",
            ) from exc

    @app.post(
        f"{API_PREFIX}/jobs/{{job_id}}/retry",
        response_model=JobAccepted,
        status_code=status.HTTP_202_ACCEPTED,
    )
    async def retry_job(job_id: str) -> JobAccepted:
        """Run a job's crawl again from scratch, with its original settings.

        A new job, never a mutation of the old one. The original record is the
        evidence of what happened and when; overwriting it would destroy the
        history an audit depends on, and a failed crawl is often the finding.

        Deliberately unrestricted by the original job's outcome: re-running a
        *successful* crawl to pick up site changes is as legitimate as retrying
        a failed one.

        Raises:
            HTTPException: `404` if there is no such job, `409` if its settings
                cannot be replayed, `400` if the URL no longer passes SSRF
                validation, `429` if no concurrency slot is free.
        """
        payload = _stored_payload(job_id)
        return _start(payload, f"{payload.base_url} (retry)")

    @app.post(f"{API_PREFIX}/jobs/{{job_id}}/reparse", response_model=JobRecord)
    async def reparse_job(job_id: str) -> JobRecord:
        """Re-run placement over a finished job under today's rules.

        Synchronous and offline. No worker thread, no concurrency slot, no
        network: the homepage body is read from this job's sidecar and the
        classified pages come from its stored result. A 27,562-page crawl
        reparses in well under a second.

        A **new** job, never a mutation of the old one, for the same reason
        retry is. The original is the evidence of what the site looked like when
        it was crawled; overwriting it would destroy the comparison that makes a
        reparse worth running.

        What this can and cannot pick up follows from what was stored. With a
        homepage sidecar the menu is re-parsed, so a `nav_tree_parser` fix
        applies. Without one — every crawl run before the sidecar existed — the
        stored menu stands and only the placement rules re-run. Breadcrumbs are
        never re-extracted: the page bodies are gone.

        Raises:
            HTTPException: `404` if there is no such job, `409` if it has no
                stored result to reparse.
        """
        try:
            stored = state.store.read_result(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND, detail=f"no job {job_id} with a result"
            ) from exc

        try:
            before = PageClassificationOutput.model_validate(stored)
        except ValidationError as exc:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail="this job's result predates the current output contract",
            ) from exc

        homepage = state.store.read_homepage(job_id)
        after = reparse_placement(before, homepage)

        record = state.store.create(
            TOOL_NAME,
            {"base_url": before.base_url},
            label=f"{before.base_url} (reparsed)",
        )
        state.store.finish(record.id, after.model_dump(mode="json"))
        # Carried over so the new job can itself be reparsed. Without this the
        # chain breaks after one hop, which is exactly when it matters.
        if homepage:
            state.store.write_homepage(record.id, homepage)
        _logger.info(
            "job_reparsed",
            extra={"source": job_id, "job_id": record.id, "menu_reparsed": bool(homepage)},
        )
        return state.store.get(record.id)

    @app.post(
        f"{API_PREFIX}/jobs/{{job_id}}/reconcile/screaming-frog",
        response_model=ReconciliationSummary,
    )
    async def reconcile_screaming_frog(job_id: str, request: Request) -> ReconciliationSummary:
        """Compare a Screaming Frog export against a finished job, and merge the gap.

        **Entirely optional.** Nothing else in the engine calls this, and a crawl
        that never sees an export behaves exactly as it always has. This is an
        extra pass an operator may run when they happen to have a Screaming Frog
        licence and a reason to cross-check.

        The body is the raw CSV, sent as `text/csv` — not `multipart/form-data`.
        FastAPI's file upload needs `python-multipart`, which is not a
        dependency of this project, and adding a package to accept one file when
        Starlette already hands over the request body would be a poor trade. The
        browser reads the file and POSTs its text.

        Synchronous and offline, like reparse: no worker thread, no concurrency
        slot, no network. Merging 250 pages into a 27,656-page crawl takes about
        1.2 seconds, nearly all of it re-running placement over the combined set.

        A **new** job when anything merges, never a mutation of the old one. The
        original is the evidence of what the crawl alone found, and the whole
        value of a reconciliation is the comparison. When nothing merges, no job
        is created and `job_id` echoes the source.

        Raises:
            HTTPException: `404` if there is no such job, `409` if its result
                predates the current output contract, `400` if the body is
                empty or is not a readable export.
        """
        body = await request.body()
        if not body.strip():
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail="empty body: POST the Screaming Frog CSV as text/csv",
            )

        try:
            stored = state.store.read_result(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND, detail=f"no job {job_id} with a result"
            ) from exc

        try:
            before = PageClassificationOutput.model_validate(stored)
        except ValidationError as exc:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail="this job's result predates the current output contract",
            ) from exc

        # Raw bytes, not text. The body may be an `.xlsx`, and decoding a
        # spreadsheet to a string turns it into mojibake that fails to parse —
        # which is how a dropped workbook once surfaced as "Is the API server
        # running?". Format detection belongs with the parser, which reads the
        # content rather than trusting a filename or a Content-Type header.
        try:
            outcome = merge_reconciled_urls(before, body)
        except ValueError as exc:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

        report = outcome.report
        target = job_id
        if outcome.merged:
            record = state.store.create(
                TOOL_NAME,
                {"base_url": before.base_url},
                label=f"{before.base_url} (+{outcome.merged} from Screaming Frog)",
            )
            state.store.finish(record.id, outcome.output.model_dump(mode="json"))
            homepage = state.store.read_homepage(job_id)
            # Carried over so the merged job can itself be reparsed later.
            if homepage:
                state.store.write_homepage(record.id, homepage)
            target = record.id

        _logger.info(
            "job_reconciled",
            extra={
                "source": job_id,
                "job_id": target,
                "merged": outcome.merged,
                "missed": len(report.missed_pages),
            },
        )
        summary = ReconciliationSummary(
            job_id=target,
            source_job_id=job_id,
            base_url=before.base_url,
            frog_rows=report.frog_rows,
            in_both=report.in_both,
            missed_pages=len(report.missed_pages),
            orphans=len(report.orphans),
            merged=outcome.merged,
            frog_reasons=dict(report.frog_reasons),
            engine_reasons=dict(report.engine_reasons),
        )

        # Saved against the *source* job, and saved with the URL lists rather
        # than the counts alone. A cross-check costs an export somebody produced
        # by hand in another tool, and it lived only in one component's state:
        # leaving the panel threw it away, and getting it back meant exporting
        # from Screaming Frog and uploading 4 MB again. The counts are the
        # headline, but the URLs are the part an analyst acts on.
        state.store.write_reconciliation(
            job_id,
            {
                "summary": summary.model_dump(mode="json"),
                "created_at": datetime.now(UTC).isoformat(),
                "missed_pages": list(report.missed_pages),
                "orphans": list(report.orphans),
                # The agreement, not only its size. Every other number on the
                # panel could be handed to someone as a list of addresses and
                # this one could not, which made it the only figure a reader had
                # to take on trust.
                "in_both": list(report.in_both_urls),
                "frog_only": [gap.model_dump(mode="json") for gap in report.frog_only],
                "engine_only": [gap.model_dump(mode="json") for gap in report.engine_only],
            },
        )
        return summary

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/reconciliation")
    def get_reconciliation(job_id: str) -> Mapping[str, object]:
        """The last Screaming Frog cross-check run against this job.

        Kept so the panel can be reopened. Returns the counts *and* the URL
        lists behind them, because "892 missed pages" is the headline and the
        892 addresses are the work.

        Raises:
            HTTPException: `404` if this job has never been cross-checked.
        """
        saved = state.store.read_reconciliation(job_id)
        if saved is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                detail=f"job {job_id} has no saved reconciliation",
            )
        return saved

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/reconciliation.csv")
    def download_reconciliation(job_id: str) -> Response:
        """The cross-check as a spreadsheet, one row per disagreement.

        The saved JSON is for the panel to redraw itself; this is for a person
        to keep. An analyst comparing two crawlers is already working in
        Screaming Frog and Excel, so the artefact they can act on is a file that
        opens there — not a modal they have to leave open.

        One flat table rather than two, with a `found_by` column, because the
        question is per URL: *which crawler saw this, and why did the other
        one not?* Splitting the two sides into separate files makes the reader
        do a join to answer it.

        Only the disagreements are listed. The URLs both crawlers found are the
        large majority and carry no finding; their count is in the summary rows
        at the top of the file.

        Raises:
            HTTPException: `404` if this job has never been cross-checked.
        """
        saved = state.store.read_reconciliation(job_id)
        if saved is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                detail=f"job {job_id} has no saved reconciliation",
            )

        summary = saved.get("summary")
        summary = summary if isinstance(summary, Mapping) else {}
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["url", "found_by", "reason", "meaning"])

        # The counts ride in the same file as pseudo-rows. A spreadsheet handed
        # to a client without them invites the reader to treat the gap lists as
        # the whole site.
        for key in ("base_url", "frog_rows", "in_both", "missed_pages", "orphans", "merged"):
            writer.writerow([summary.get(key, ""), "summary", key, ""])

        for side, rows in (
            ("screaming_frog_only", saved.get("frog_only")),
            ("rankuno_only", saved.get("engine_only")),
        ):
            for row in rows if isinstance(rows, list) else []:
                if not isinstance(row, Mapping):
                    continue
                reason = str(row.get("reason", ""))
                writer.writerow([row.get("url", ""), side, reason, GAP_MEANINGS.get(reason, "")])

        stamp = str(saved.get("created_at", ""))[:10]
        name = f"cross-check-{job_id[:8]}-{stamp or 'undated'}.csv"
        return _csv_response(buffer.getvalue(), name)

    @app.post(
        f"{API_PREFIX}/jobs/{{job_id}}/performance/gsc",
        response_model=PerformanceSummary,
    )
    async def upload_gsc(job_id: str, request: Request) -> PerformanceSummary:
        """Attach a Search Console page export to a finished crawl.

        The body is the raw file, sent as `application/octet-stream` — not
        `multipart/form-data`, for the same reason as the Screaming Frog
        endpoint: FastAPI's file upload needs `python-multipart`, and adding a
        dependency to accept one file that Starlette already hands over as a
        request body is a poor trade. The browser passes the `File` straight to
        `fetch`, which streams it.

        Whatever Search Console produced is accepted — the ZIP from Export →
        CSV, the workbook from Export → Excel, or a bare CSV somebody unpacked.
        The parser picks the pages tab by content, because the archive is
        written in the account's display language.

        Synchronous and offline, like reparse and reconcile: no worker thread,
        no concurrency slot, no network. Resolving 1,000 rows against a
        12,787-page crawl and rolling them up takes under a second.

        **Nothing about the crawl changes.** No new job, no mutation of the
        result. The report is a sidecar the crawl knows nothing about, so a job
        that never sees an export behaves exactly as it always has.

        Re-uploading replaces the previous report. That is how somebody corrects
        a wrong date range or the wrong property, and keeping the superseded one
        would leave two reports with no way to tell which is being looked at.

        Raises:
            HTTPException: `404` if there is no such job, `409` if its result
                predates the current output contract, `400` if the body is
                empty, oversized, or not a readable export.
        """
        body = await request.body()
        if not body.strip():
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail="empty body: POST the Search Console export as the request body",
            )
        if len(body) > MAX_PERFORMANCE_UPLOAD:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"upload is {len(body) // (1024 * 1024)} MB, over the "
                    f"{MAX_PERFORMANCE_UPLOAD // (1024 * 1024)} MB limit"
                ),
            )

        try:
            stored = state.store.read_result(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND, detail=f"no job {job_id} with a result"
            ) from exc

        try:
            crawl = PageClassificationOutput.model_validate(stored)
        except ValidationError as exc:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail="this job's result predates the current output contract",
            ) from exc

        # Raw bytes, never decoded first. The body is usually a ZIP, and
        # decoding an archive to a string turns it into mojibake that parses as
        # nothing — the failure the Screaming Frog endpoint already shipped once
        # and surfaced as "Is the API server running?".
        try:
            export = load_gsc_export(body)
        except ValueError as exc:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

        index = UrlResolutionIndex(crawl.pages)
        metrics = merge_page_metrics(index, export.rows)
        rollup = rollup_of(index, metrics)
        opportunities = score_opportunities(index, metrics)

        summary = PerformanceSummary(
            job_id=job_id,
            base_url=crawl.base_url,
            source_name=export.source_name,
            rows=len(export.rows),
            skipped_rows=export.skipped_rows,
            matched=rollup.gsc_resolution.matched_count,
            match_rate_pct=rollup.gsc_resolution.match_rate_pct,
            is_reliable=rollup.gsc_resolution.is_reliable,
            pages_with_data=rollup.site.pages_with_data,
            pages=rollup.site.pages,
            unmatched=unmatched_groups(metrics),
            rollup=rollup,
            opportunities=opportunities,
        )
        _logger.info(
            "performance_attached",
            extra={
                "job_id": job_id,
                "rows": len(export.rows),
                "match_rate_pct": summary.match_rate_pct,
                "source": export.source_name,
            },
        )
        state.store.write_performance(
            job_id,
            {
                "summary": summary.model_dump(mode="json"),
                "created_at": datetime.now(UTC).isoformat(),
                # Every unresolved row, not just the grouping. "585 rows reached
                # no page" is the headline and the 585 addresses are what an
                # analyst checks it against — and re-deriving them would mean
                # asking for the export again.
                "matched_rows": _matched_rows(index, metrics),
                # Every unresolved row, not just the grouping. "585 rows reached
                # no page" is the headline and the 585 addresses are what an
                # analyst checks it against.
                "unmatched_rows": [
                    {
                        "url": row.url,
                        "clicks": row.clicks,
                        "impressions": row.impressions,
                        "reason": reason,
                    }
                    for row, reason in _unmatched_rows(metrics)
                ],
            },
        )
        return summary

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/performance")
    def get_performance(job_id: str) -> Mapping[str, object]:
        """The last Search Console report attached to this job.

        Kept so the panel redraws itself without asking for the file again — the
        export took a person a trip to another product to obtain.

        Raises:
            HTTPException: `404` if no export has been attached to this job.
        """
        saved = state.store.read_performance(job_id)
        if saved is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                detail=f"job {job_id} has no attached Search Console data",
            )
        return saved

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/opportunities.csv")
    def download_opportunities(job_id: str) -> Response:
        """The recommendations as a spreadsheet, one row per finding.

        The panel is for reading; this is the artefact that gets assigned. It
        carries the plain-language reason rather than the enum, because the
        person who acts on the row is usually not the person who ran the crawl.

        The kinds that were **skipped** ride in the same file as pseudo-rows. A
        list of recommendations handed over without them invites the reader to
        conclude the site has no orphans, when the truth is that this crawl
        could not tell.

        Raises:
            HTTPException: `404` if no export has been attached to this job.
        """
        saved = state.store.read_performance(job_id)
        if saved is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                detail=f"job {job_id} has no attached Search Console data",
            )
        summary = saved.get("summary")
        summary = summary if isinstance(summary, Mapping) else {}
        report = summary.get("opportunities")
        report = report if isinstance(report, Mapping) else {}

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            [
                "severity",
                "kind",
                "score",
                "url",
                "section",
                "clicks",
                "impressions",
                "position",
                "reason",
            ]
        )
        rows = report.get("opportunities")
        for row in rows if isinstance(rows, list) else []:
            if not isinstance(row, Mapping):
                continue
            section = row.get("section")
            writer.writerow(
                [
                    # First column, so a critical row is visible before anybody
                    # widens anything or sorts anything.
                    row.get("severity", ""),
                    row.get("kind", ""),
                    row.get("score", ""),
                    row.get("url", ""),
                    " > ".join(str(part) for part in section) if isinstance(section, list) else "",
                    row.get("clicks", ""),
                    row.get("impressions", ""),
                    row.get("position", ""),
                    row.get("reason", ""),
                ]
            )

        skipped = report.get("skipped")
        for kind, gap in (skipped if isinstance(skipped, Mapping) else {}).items():
            writer.writerow(["", kind, "", "", "", "", "", "", f"not evaluated: {gap}"])

        stamp = str(saved.get("created_at", ""))[:10]
        name = f"opportunities-{job_id[:8]}-{stamp or 'undated'}.csv"
        return _csv_response(buffer.getvalue(), name)

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/matched.csv")
    def download_matched(job_id: str) -> Response:
        """The pages the export did reach, with the crawl's columns beside them.

        The other half of `unmatched.csv`, and the more useful half: this is the
        joined dataset. Search Console gives clicks against a URL and knows
        nothing about where that URL sits; the crawl knows the navigation
        section, the page type and the inbound-link count and nothing about
        traffic. Neither file answers "which section earns" on its own.

        Sorted by clicks. One row per crawled page, not per export row — several
        Google URLs can name one page, and they were summed on the way in.

        Raises:
            HTTPException: `404` if no export has been attached, or `409` if the
                attached report predates this download and holds no page rows.
        """
        saved = state.store.read_performance(job_id)
        if saved is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                detail=f"job {job_id} has no attached Search Console data",
            )
        rows = saved.get("matched_rows")
        if not isinstance(rows, list):
            # An older report. Saying so beats an empty file, which reads as
            # "nothing matched" when the truth is "this was saved before the
            # download existed".
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=(
                    "this report was saved before per-page rows were kept. "
                    "Upload the export again to produce it."
                ),
            )

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        columns = [
            "url",
            "section",
            "clicks",
            "impressions",
            "ctr",
            "position",
            "page_type",
            "hierarchy_level",
            "inbound_internal_links",
        ]
        writer.writerow(columns)
        for row in rows:
            if isinstance(row, Mapping):
                writer.writerow([row.get(column, "") for column in columns])

        stamp = str(saved.get("created_at", ""))[:10]
        name = f"matched-{job_id[:8]}-{stamp or 'undated'}.csv"
        return _csv_response(buffer.getvalue(), name)

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/unmatched.csv")
    def download_unmatched(job_id: str) -> Response:
        """Every export row that reached no crawled page.

        The evidence behind the match rate. "41.5% matched" is a number an
        analyst has to take on trust; this is the other 58.5%, one row each,
        with the clicks they carry and the reason they did not land — so the
        rate can be checked rather than believed, and so the URLs themselves can
        be acted on.

        Sorted by clicks, because the gap is not evenly distributed and the
        first rows are usually the whole story.

        The group totals ride at the top as summary rows, for the same reason
        they do in the cross-check download: a list of addresses handed over
        without them invites the reader to count the rows and think that is the
        finding.

        Raises:
            HTTPException: `404` if no export has been attached to this job.
        """
        saved = state.store.read_performance(job_id)
        if saved is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                detail=f"job {job_id} has no attached Search Console data",
            )
        rows = saved.get("unmatched_rows")
        if not isinstance(rows, list):
            # The same guard `matched.csv` carries, and for the same reason: a
            # header-only file reads as "every row matched", which is the
            # opposite of what an older report means. Shipped on one endpoint
            # and not the other, and found by clicking the link.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=(
                    "this report was saved before the unmatched rows were kept. "
                    "Upload the export again to produce it."
                ),
            )
        summary = saved.get("summary")
        summary = summary if isinstance(summary, Mapping) else {}

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["url", "host", "reason", "meaning", "clicks", "impressions"])

        groups = summary.get("unmatched")
        for group in groups if isinstance(groups, list) else []:
            if not isinstance(group, Mapping):
                continue
            reason = str(group.get("reason", ""))
            writer.writerow(
                [
                    f"{group.get('urls', 0)} URLs",
                    group.get("host", ""),
                    reason,
                    "group total",
                    group.get("clicks", 0),
                    group.get("impressions", 0),
                ]
            )

        for row in rows:
            if not isinstance(row, Mapping):
                continue
            url = str(row.get("url", ""))
            parts = safe_split(url)
            reason = str(row.get("reason", ""))
            writer.writerow(
                [
                    url,
                    site_host(parts.netloc) if parts is not None else "",
                    reason,
                    UNMATCHED_MEANINGS.get(reason, ""),
                    row.get("clicks", 0),
                    row.get("impressions", 0),
                ]
            )

        stamp = str(saved.get("created_at", ""))[:10]
        name = f"unmatched-{job_id[:8]}-{stamp or 'undated'}.csv"
        return _csv_response(buffer.getvalue(), name)

    @app.get(f"{API_PREFIX}/jobs/{{job_id}}/reconciliation.xlsx")
    def download_reconciliation_workbook(job_id: str) -> Response:
        """The cross-check as a workbook, one sheet per question.

        The same data as `reconciliation.csv`, which stays for anything already
        linking to it. This is the one to hand somebody: a real gep.com
        cross-check is 17,640 rows, and in a single sheet the 14 pages the crawl
        actually missed sit below 16,000 rows of differences that are explained
        and need no action.

        Five sheets, smallest first:

        * **Summary** — the counts, and every reason with what it means.
        * **Missed pages** — live, in-scope pages the crawl never reached. The
          engine's own defect, and already merged.
        * **Orphans** — published pages no internal link reaches. Found only by
          this engine, and left alone: their absence from the other crawl *is*
          the finding.
        * **Screaming Frog only** / **Rankuno only** — every remaining
          difference with the reason it is not a defect.

        Raises:
            HTTPException: `404` if this job has never been cross-checked.
        """
        saved = state.store.read_reconciliation(job_id)
        if saved is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                detail=f"job {job_id} has no saved cross-check",
            )
        summary = saved.get("summary")
        summary = summary if isinstance(summary, Mapping) else {}

        # One bucket per reason, across both sides. The two vocabularies are
        # disjoint — `FrogGapReason` and `EngineGapReason` share no member — so a
        # reason names its side without needing a prefix, and the dedicated
        # `missed_pages` and `orphans` lists are exactly the `MISSED_PAGE` and
        # `SITEMAP_ORPHAN` buckets rather than anything extra.
        buckets: dict[str, list[str]] = {}
        sides: dict[str, str] = {}
        for key, side in (("frog_only", "Screaming Frog"), ("engine_only", "Rankuno")):
            found = saved.get(key)
            for row in found if isinstance(found, list) else []:
                if not isinstance(row, Mapping):
                    continue
                reason = str(row.get("reason", ""))
                buckets.setdefault(reason, []).append(str(row.get("url", "")))
                sides[reason] = side

        # The two findings first, then the rest largest first. On a real
        # cross-check `MEDIA_URL` is 16,162 of 16,337 rows, so by size alone the
        # 15 pages the crawl missed would sit at the far end of the workbook.
        def rank(reason: str) -> tuple[int, int]:
            lead = {"MISSED_PAGE": 0, "SITEMAP_ORPHAN": 1}.get(reason, 2)
            return (lead, -len(buckets[reason]))

        ordered = sorted(buckets, key=rank)

        counts: list[list[object]] = [
            ["Cross-checked", summary.get("base_url", "")],
            ["Run at", str(saved.get("created_at", ""))[:19]],
            ["Rows in the Screaming Frog export", summary.get("frog_rows", "")],
            ["Found by both crawlers", summary.get("in_both", "")],
            ["Pages this crawl missed (merged)", summary.get("missed_pages", "")],
            ["Orphans only this crawl found", summary.get("orphans", "")],
            ["URLs merged into a new job", summary.get("merged", "")],
            [],
            ["Why Screaming Frog saw a URL this crawl did not", "URLs", "Meaning"],
        ]

        def tally(key: str) -> list[list[object]]:
            found = summary.get(key)
            return [
                [reason, count, GAP_MEANINGS.get(str(reason), "")]
                for reason, count in (found if isinstance(found, Mapping) else {}).items()
            ]

        counts.extend(tally("frog_reasons"))
        counts.append([])
        counts.append(["Why this crawl saw a URL Screaming Frog did not", "URLs", "Meaning"])
        counts.extend(tally("engine_reasons"))

        # A contents page. With a dozen sheets the reader needs to know which
        # one to open, and this is also where each reason's meaning lives now
        # that it is not repeated down every row of its own sheet.
        counts.append([])
        counts.append(["Sheet", "Found by", "URLs", "What it means"])
        counts.extend(
            [
                SHEET_TITLES.get(reason, reason),
                sides.get(reason, ""),
                len(buckets[reason]),
                GAP_MEANINGS.get(reason, ""),
            ]
            for reason in ordered
        )

        sheets: list[Sheet] = [("Summary", ["Measure", "Value", "Meaning"], counts)]
        sheets.extend(
            # One column, not three. The reason is the sheet and the meaning is
            # on the contents page, so carrying both down every row repeats one
            # value 16,162 times — which is what the flat version looked like
            # and why it was unreadable.
            (SHEET_TITLES.get(reason, reason), ["url"], [[url] for url in buckets[reason]])
            for reason in ordered
        )
        stamp = str(saved.get("created_at", ""))[:10]
        name = f"cross-check-{job_id[:8]}-{stamp or 'undated'}.xlsx"
        return _workbook_response(sheets, name)

    @app.post(f"{API_PREFIX}/jobs/{{job_id}}/cancel", response_model=JobRecord)
    async def cancel_job(job_id: str) -> JobRecord:
        """Abandon a job and give its concurrency slot back.

        **This releases the slot; it does not stop the crawl.** The work runs on
        a worker thread via `asyncio.to_thread`, and a Python thread cannot be
        killed from outside — cancelling the awaiting task does not reach into
        it either. The thread keeps fetching until it finishes or the process
        exits, and its result is discarded when it does.

        That is a weaker guarantee than the button implies, and it is still
        worth having. The failure this exists for is a crawl wedged in network
        I/O for hours: two stripe.com jobs held two of three slots for sixteen
        hours on one workstation, so every new crawl was refused by a server
        that was, for practical purposes, doing nothing. Releasing the slot
        restores the ability to work; waiting for the thread does not.

        A real stop needs a cancellation flag the crawl checks between fetches.
        That does not exist yet — see the build log for this cycle.

        Raises:
            HTTPException: `404` if there is no such job, `409` if it has
                already reached a terminal state.
        """
        try:
            record = state.store.get(job_id)
        except JobNotFoundError as exc:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"no job {job_id}") from exc
        if record.status not in {JobStatus.QUEUED, JobStatus.RUNNING}:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=f"job is {record.status.value} and has already finished",
            )

        state.release(job_id)
        updated = state.store.mark_failed(
            job_id,
            "cancelled by operator — the crawl thread may still be running until it "
            "finishes or the server restarts",
        )
        _logger.warning("job_cancelled", extra={"job_id": job_id})
        return updated

    @app.post(
        f"{API_PREFIX}/jobs/{{job_id}}/resume",
        response_model=JobAccepted,
        status_code=status.HTTP_202_ACCEPTED,
    )
    async def resume_job(job_id: str) -> JobAccepted:
        """Crawl the URLs an interrupted job discovered but never fetched.

        A separate job producing a separate result — **not** a merge into the
        original. Merging is not a matter of appending pages: inbound link
        counts, orphan flags and navigation coverage are properties of the whole
        graph, and a checkpoint stores URLs only. Classifying the remainder
        against a graph missing the pages already crawled would produce wrong
        in-degrees and wrong orphan flags, and orphan detection is a headline
        finding in these reports.

        Refused unless the original crawl genuinely stopped early. A crawl that
        ran to completion still shows more URLs discovered than fetched — a
        sitemap lists pages no link reaches, and faceted filters are declined on
        purpose — so `discovered > fetched` is normal and is not unfinished
        work. `truncated` and `stopped_reason` are what distinguish the two.

        Raises:
            HTTPException: `404` if there is no such job or it saved no partial
                work, `409` if it did not stop early, is still running, or has
                nothing left to fetch, and the codes `retry` can raise.
        """
        payload = _stored_payload(job_id)
        record = state.store.get(job_id)

        if not record.is_terminal:
            # Its checkpoint is still moving, so any delta read now is stale
            # before the resumed crawl starts.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=f"job {job_id} is still {record.status.value}",
            )

        checkpoint = state.store.read_checkpoint(job_id)
        if checkpoint is None:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND, detail=f"job {job_id} saved no partial work"
            )

        unfetched = checkpoint.get("unfetched")
        if not isinstance(unfetched, list):
            # Written before checkpoints recorded this. The discovered set alone
            # cannot tell fetched from unfetched, so resuming would re-fetch
            # everything — which is what `retry` is, said honestly.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=(
                    f"job {job_id} predates resumable checkpoints and cannot be "
                    "resumed; retry it instead"
                ),
            )

        remaining = tuple(url for url in unfetched if isinstance(url, str))
        if not remaining:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=f"job {job_id} fetched every URL it discovered",
            )

        # The other half, and the half that was missing. `seed_urls` only adds
        # to the frontier; without an exclusion the crawl still begins at the
        # site root, follows every link out of it and re-fetches the whole site,
        # with the seeds appended to a run that was never a resume. Observed on
        # gep.com: a resume advertising "+2,940" rediscovered 5,311 URLs and
        # started fetching from zero.
        #
        # Derived rather than stored: a checkpoint records everything it
        # discovered and what it had not yet fetched, so the difference is
        # exactly what the interrupted run completed.
        discovered = checkpoint.get("urls")
        if isinstance(discovered, list):
            outstanding = set(remaining)
            already = tuple(
                url for url in discovered if isinstance(url, str) and url not in outstanding
            )
        else:
            # A checkpoint old enough to lack the discovered set. Resuming still
            # beats refusing — the seeds are real work — but it degrades to the
            # old behaviour, so it is logged rather than passed over in silence.
            already = ()
            _logger.warning("resume_without_exclusion", extra={"job_id": job_id})

        resumed = payload.model_copy(update={"seed_urls": remaining, "exclude_urls": already})
        _logger.info(
            "job_resumed",
            extra={"source": job_id, "seeds": len(remaining), "excluded": len(already)},
        )
        return _start(resumed, f"{payload.base_url} (resumed +{len(remaining):,})")

    return app


def _checkpoint_as_output(checkpoint: Mapping[str, object]) -> PageClassificationOutput:
    """Rebuild a renderable output from saved URLs.

    The classifications are placeholders, and `stopped_reason` says so in the one
    place every consumer already reads — so a recovered view cannot be mistaken
    for a completed crawl.
    """
    base_url = str(checkpoint.get("base_url") or "")
    raw = checkpoint.get("urls")
    urls = [str(item) for item in raw] if isinstance(raw, list) else []

    return PageClassificationOutput(
        base_url=base_url,
        site_profile=SiteProfile(),
        weight_profile=WeightProfileReport.for_site(SiteProfile()),
        discovery=DiscoveryReport(
            base_url=base_url,
            total_urls=len(urls),
            stopped_reason=(
                "recovered from a checkpoint — these URLs were discovered before the "
                "crawl was interrupted, and none of them were classified"
            ),
        ),
        summary=CrawlSummary(pages_classified=len(urls), unknown_pages=len(urls)),
        pages=tuple(_placeholder_profile(url) for url in urls),
    )


def _placeholder_profile(url: str) -> FullPageIntelligenceProfile:
    """An unclassified page, carrying only what a checkpoint actually knows."""
    return FullPageIntelligenceProfile(
        url=url,
        canonical_url=url,
        normalized_path=url,
        hierarchy_level=HierarchyLevel.L3_LEAF_PAGE,
        primary_page_type=PrimaryPageType.UNKNOWN,
        depth_from_l0=1,
        search_intent=SearchIntent.INFORMATIONAL,
        signals_evaluated=(
            SignalScore(
                source=SignalSource.SITEMAP_INDEX,
                suggested_level=HierarchyLevel.L3_LEAF_PAGE,
                suggested_page_type=PrimaryPageType.UNKNOWN,
                confidence=0.0,
                notes="discovered before the crawl was interrupted; never classified",
            ),
        ),
        final_confidence_score=0.0,
        consensus_method=ConsensusMethod.LAYER0_FAST_PATH,
    )


class ServerConfig(StrictModel):
    """Bind settings for `serve()`."""

    host: str = "127.0.0.1"
    port: int = Field(default=8000, gt=0, le=65535)
    jobs_root: str = ".jobs"


def serve(config: ServerConfig | None = None) -> None:  # pragma: no cover - process entry point
    """Run the server.

    Binds loopback by default and that default should not be changed casually:
    there is no authentication, and the server fetches arbitrary URLs on
    request. Exposed on a routable interface it is an open proxy.
    """
    import uvicorn

    settings = config or ServerConfig()
    uvicorn.run(
        create_app(jobs_root=settings.jobs_root),
        host=settings.host,
        port=settings.port,
        log_config=None,
    )


if __name__ == "__main__":  # pragma: no cover
    serve()
